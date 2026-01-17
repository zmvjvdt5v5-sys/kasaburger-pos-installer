# KasaBurger API Server - v1.0.3
# Health check MUST be first - before any imports that could fail

from fastapi import FastAPI
from datetime import datetime, timezone, timedelta
import os
from starlette.middleware.cors import CORSMiddleware


# Create app immediately
app = FastAPI(title="KasaBurger API", version="1.0.3")

# CORS Configuration - Allow all origins if not specified
cors_origins_env = os.getenv("CORS_ORIGINS", "*")
if cors_origins_env == "*":
    origins = ["*"]
else:
    origins = [origin.strip() for origin in cors_origins_env.split(",") if origin.strip()]
    if not origins:
        origins = ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "PATCH"],
    allow_headers=["*"],
    expose_headers=["*"],
    max_age=3600,
)

# Health check - FIRST route, no dependencies
@app.get("/health")
def health_check():
    return {"status": "healthy", "version": "1.0.3", "timestamp": datetime.now(timezone.utc).isoformat()}

# Root endpoint - also returns healthy status for proxy compatibility
@app.get("/")
def root():
    return {"status": "healthy", "version": "1.0.3", "timestamp": datetime.now(timezone.utc).isoformat()}

# Now load the rest of the application
try:
    from fastapi import APIRouter, HTTPException, Depends, status, Request, Body
    from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
    from fastapi.responses import StreamingResponse, JSONResponse
    from dotenv import load_dotenv
    from starlette.middleware.base import BaseHTTPMiddleware
    from motor.motor_asyncio import AsyncIOMotorClient
    import logging
    from pathlib import Path
    from pydantic import BaseModel, Field, EmailStr
    from typing import List, Optional, Dict
    import uuid
    import jwt
    import bcrypt
    import io
    import httpx
    import time
    from collections import defaultdict
    import hashlib
    import cloudinary
    import cloudinary.utils
    import cloudinary.uploader

    # NOT: Cloudinary config'i load_dotenv'den sonra yapılacak (satır ~180)

    # ==================== SECURITY CONFIGURATION ====================
    
    # Rate limiting storage (in-memory, resets on restart)
    rate_limit_storage: Dict[str, List[float]] = defaultdict(list)
    failed_login_attempts: Dict[str, List[float]] = defaultdict(list)
    blocked_ips: Dict[str, float] = {}
    
    # Security settings
    RATE_LIMIT_REQUESTS = 100  # Max requests per window
    RATE_LIMIT_WINDOW = 60  # Window in seconds
    LOGIN_ATTEMPT_LIMIT = 50  # Max failed login attempts
    LOGIN_BLOCK_DURATION = 10  # Block duration in seconds (10 saniye)
    BULK_REQUEST_LIMIT = 10  # Max bulk requests (like get all products) per minute
    
    # Allowed origins (production domains)
    # ALLOWED_ORIGINS = [
    #     "https://burger-pos-1.preview.emergentagent.com",
    #     "https://kasaburger.net.tr",
    #     "https://www.kasaburger.net.tr",
    #     "http://localhost:3000",
    # ]

    # ==================== SECURITY MIDDLEWARE ====================
    
    class SecurityMiddleware(BaseHTTPMiddleware):
        async def dispatch(self, request: Request, call_next):
            client_ip = request.client.host if request.client else "unknown"
            path = request.url.path
            method = request.method
            
            # Skip security for health checks
            if path in ["/health", "/", "/api/health"]:
                return await call_next(request)
            
            # Check if IP is blocked
            if client_ip in blocked_ips:
                if time.time() < blocked_ips[client_ip]:
                    return JSONResponse(
                        status_code=429,
                        content={"detail": "IP adresi geçici olarak engellendi. Lütfen daha sonra tekrar deneyin."}
                    )
                else:
                    del blocked_ips[client_ip]
            
            # Rate limiting
            current_time = time.time()
            rate_limit_storage[client_ip] = [
                t for t in rate_limit_storage[client_ip] 
                if current_time - t < RATE_LIMIT_WINDOW
            ]
            
            if len(rate_limit_storage[client_ip]) >= RATE_LIMIT_REQUESTS:
                # Log suspicious activity
                logging.warning(f"Rate limit exceeded for IP: {client_ip}, Path: {path}")
                return JSONResponse(
                    status_code=429,
                    content={"detail": "Çok fazla istek gönderildi. Lütfen bir dakika bekleyin."}
                )
            
            rate_limit_storage[client_ip].append(current_time)
            
            # Extra protection for bulk data endpoints
            bulk_endpoints = ["/api/products", "/api/dealers", "/api/orders", "/api/invoices", "/api/payments"]
            if path in bulk_endpoints and method == "GET":
                bulk_key = f"{client_ip}_bulk"
                rate_limit_storage[bulk_key] = [
                    t for t in rate_limit_storage[bulk_key] 
                    if current_time - t < 60
                ]
                if len(rate_limit_storage[bulk_key]) >= BULK_REQUEST_LIMIT:
                    logging.warning(f"Bulk request limit exceeded for IP: {client_ip}")
                    return JSONResponse(
                        status_code=429,
                        content={"detail": "Toplu veri istekleri sınırlandırıldı. Lütfen bekleyin."}
                    )
                rate_limit_storage[bulk_key].append(current_time)
            
            # Add security headers to response
            response = await call_next(request)
            response.headers["X-Content-Type-Options"] = "nosniff"
            response.headers["X-Frame-Options"] = "DENY"
            response.headers["X-XSS-Protection"] = "1; mode=block"
            response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
            response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
            
            return response
    
    # Add security middleware
    app.add_middleware(SecurityMiddleware)

    # Add CORS with restricted origins
    # app.add_middleware(
    #     CORSMiddleware,
    #     allow_credentials=True,
    #     allow_origins=ALLOWED_ORIGINS,
    #     allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    #     allow_headers=["Authorization", "Content-Type", "X-Requested-With"],
    # )

    ROOT_DIR = Path(__file__).parent
    load_dotenv(ROOT_DIR / '.env')

    # Cloudinary Configuration (load_dotenv'den sonra yapılmalı)
    cloudinary.config(
        cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
        api_key=os.getenv("CLOUDINARY_API_KEY"),
        api_secret=os.getenv("CLOUDINARY_API_SECRET"),
        secure=True
    )

    # Şube (Branch) Configuration
    BRANCH_ID = os.environ.get('BRANCH_ID', 'main')
    BRANCH_NAME = os.environ.get('BRANCH_NAME', 'Merkez')
    CENTRAL_SERVER_URL = os.environ.get('CENTRAL_SERVER_URL', '')
    CENTRAL_API_KEY = os.environ.get('CENTRAL_API_KEY', '')

    # MongoDB connection with proper settings for production
    mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
    db_name = os.environ.get('DB_NAME', 'kasaburger_db')
    
    # Connection settings optimized for cloud deployment
    client = AsyncIOMotorClient(
        mongo_url,
        serverSelectionTimeoutMS=5000,
        connectTimeoutMS=10000,
        socketTimeoutMS=20000,
        maxPoolSize=10,
        minPoolSize=1,
        retryWrites=True,
        retryReads=True
    )
    db = client[db_name]

    # JWT Configuration
    JWT_SECRET = os.environ.get('JWT_SECRET', 'kasaburger_jwt_secret_production_2024_secure_key')
    JWT_ALGORITHM = "HS256"
    JWT_EXPIRATION_HOURS = 24

    # Bizim Hesap API Configuration
    BIZIMHESAP_API_KEY = os.environ.get('BIZIMHESAP_API_KEY', '')
    BIZIMHESAP_API_URL = os.environ.get('BIZIMHESAP_API_URL', 'https://bizimhesap.com/api/b2b')

    # iyzico Payment Gateway Configuration
    IYZICO_API_KEY = os.environ.get('IYZICO_API_KEY', '')
    IYZICO_SECRET_KEY = os.environ.get('IYZICO_SECRET_KEY', '')
    IYZICO_BASE_URL = os.environ.get('IYZICO_BASE_URL', 'https://sandbox-api.iyzipay.com')
    
    # Import iyzipay
    try:
        import iyzipay
        iyzico_available = True
        logging.info("iyzico payment gateway initialized")
    except ImportError:
        iyzico_available = False
        logging.warning("iyzipay not installed - payment gateway disabled")

    api_router = APIRouter(prefix="/api")
    security = HTTPBearer()

    # Lazy imports for optional features
    reportlab_available = False
    openpyxl_available = False

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        reportlab_available = True
    except ImportError:
        pass

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        openpyxl_available = True
    except ImportError:
        pass

    # ==================== MODELS ====================

    class UserCreate(BaseModel):
        email: EmailStr
        password: str
        name: str
        role: str = "admin"

    class UserLogin(BaseModel):
        email: EmailStr
        password: str
        captcha_answer: Optional[int] = None  # Captcha yanıtı

    class TwoFactorVerify(BaseModel):
        email: EmailStr
        code: str

    class UserResponse(BaseModel):
        id: str
        email: str
        name: str
        role: str
        created_at: str
        two_factor_enabled: Optional[bool] = False

    class TokenResponse(BaseModel):
        access_token: str
        token_type: str = "bearer"
        user: UserResponse
        requires_2fa: Optional[bool] = False

    # 2FA codes storage (in-memory, production'da Redis kullanılmalı)
    two_factor_codes: Dict[str, dict] = {}
    
    # Captcha storage
    captcha_storage: Dict[str, dict] = {}
    
    # Audit log storage
    audit_logs: List[dict] = []
    
    # Password policy settings
    PASSWORD_MIN_LENGTH = 6
    PASSWORD_REQUIRE_UPPERCASE = False
    PASSWORD_REQUIRE_LOWERCASE = False
    PASSWORD_REQUIRE_NUMBER = False
    PASSWORD_REQUIRE_SPECIAL = False

    def validate_password_strength(password: str) -> tuple[bool, str]:
        """Şifre güçlülük kontrolü"""
        if len(password) < PASSWORD_MIN_LENGTH:
            return False, f"Şifre en az {PASSWORD_MIN_LENGTH} karakter olmalı"
        if PASSWORD_REQUIRE_UPPERCASE and not any(c.isupper() for c in password):
            return False, "Şifre en az bir büyük harf içermeli"
        if PASSWORD_REQUIRE_LOWERCASE and not any(c.islower() for c in password):
            return False, "Şifre en az bir küçük harf içermeli"
        if PASSWORD_REQUIRE_NUMBER and not any(c.isdigit() for c in password):
            return False, "Şifre en az bir rakam içermeli"
        if PASSWORD_REQUIRE_SPECIAL and not any(c in "!@#$%^&*()_+-=[]{}|;:,.<>?" for c in password):
            return False, "Şifre en az bir özel karakter içermeli"
        return True, "OK"

    def generate_2fa_code() -> str:
        """6 haneli 2FA kodu üret"""
        import random
        return str(random.randint(100000, 999999))

    def generate_captcha() -> dict:
        """Basit matematik captcha üret"""
        import random
        num1 = random.randint(1, 10)
        num2 = random.randint(1, 10)
        operation = random.choice(['+', '-'])
        if operation == '+':
            answer = num1 + num2
            question = f"{num1} + {num2} = ?"
        else:
            # Negatif sonuç olmasın
            if num1 < num2:
                num1, num2 = num2, num1
            answer = num1 - num2
            question = f"{num1} - {num2} = ?"
        
        captcha_id = str(uuid.uuid4())
        return {"id": captcha_id, "question": question, "answer": answer}

    def log_audit(action: str, user_id: str, user_email: str, ip: str, details: str = ""):
        """Audit log kaydı"""
        audit_logs.append({
            "id": str(uuid.uuid4()),
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "action": action,
            "user_id": user_id,
            "user_email": user_email,
            "ip_address": ip,
            "details": details
        })
        # Son 1000 kaydı tut
        if len(audit_logs) > 1000:
            audit_logs.pop(0)

    class ProductCreate(BaseModel):
        name: str
        code: str
        unit: str = "kg"
        base_price: float
        description: Optional[str] = ""

    class ProductResponse(BaseModel):
        id: str
        name: str
        code: str
        unit: str
        base_price: float
        description: str
        created_at: str

    class MaterialCreate(BaseModel):
        name: str
        code: str
        unit: str
        stock_quantity: float = 0
        min_stock: float = 0
        unit_price: float = 0

    class MaterialResponse(BaseModel):
        id: str
        name: str
        code: str
        unit: str
        stock_quantity: float
        min_stock: float
        unit_price: float
        created_at: str

    class RecipeIngredient(BaseModel):
        material_id: str
        material_name: str
        quantity: float
        unit: str

    class RecipeCreate(BaseModel):
        product_id: str
        product_name: str
        ingredients: List[RecipeIngredient]
        yield_quantity: float
        yield_unit: str = "kg"
        notes: Optional[str] = ""

    class RecipeResponse(BaseModel):
        id: str
        product_id: str
        product_name: str
        ingredients: List[RecipeIngredient]
        yield_quantity: float
        yield_unit: str
        notes: str
        created_at: str

    class ProductionCreate(BaseModel):
        recipe_id: str
        product_name: str
        quantity: float
        planned_date: str
        notes: Optional[str] = ""

    class ProductionResponse(BaseModel):
        id: str
        recipe_id: str
        product_name: str
        quantity: float
        planned_date: str
        status: str
        notes: str
        created_at: str
        completed_at: Optional[str] = None

    class DealerPricing(BaseModel):
        product_id: str
        product_name: str
        special_price: float

    class DealerCreate(BaseModel):
        name: str
        code: str
        contact_person: str
        phone: str
        email: Optional[str] = ""
        address: str
        tax_number: Optional[str] = ""
        pricing: Optional[List[DealerPricing]] = []
        password: Optional[str] = None  # Admin tarafından belirlenen şifre
        credit_limit: Optional[float] = 0  # Kredi limiti (0 = limitsiz)

    class DealerResponse(BaseModel):
        id: str
        name: str
        code: str
        contact_person: str
        phone: str
        email: str
        address: str
        tax_number: str
        pricing: List[DealerPricing]
        balance: float
        credit_limit: float = 0
        created_at: str

    class OrderItem(BaseModel):
        product_id: str
        product_name: str
        quantity: float
        unit_price: float
        total: float

    class OrderCreate(BaseModel):
        dealer_id: str
        dealer_name: str
        items: List[OrderItem]
        delivery_date: str
        notes: Optional[str] = ""

    class OrderResponse(BaseModel):
        id: str
        order_number: str
        dealer_id: str
        dealer_name: str
        items: List[OrderItem]
        subtotal: float
        tax_amount: float
        total: float
        delivery_date: str
        status: str
        notes: str
        created_at: str

    class InvoiceCreate(BaseModel):
        order_id: str
        dealer_id: str
        dealer_name: str
        items: List[OrderItem]
        subtotal: float
        tax_rate: float = 20
        tax_amount: float
        total: float
        due_date: str

    class InvoiceResponse(BaseModel):
        id: str
        invoice_number: str
        order_id: str
        dealer_id: str
        dealer_name: str
        items: List[OrderItem]
        subtotal: float
        tax_rate: float = 20.0
        tax_amount: float
        total: float
        due_date: Optional[str] = None
        status: str
        created_at: str
        paid_at: Optional[str] = None
        bizimhesap_guid: Optional[str] = None
        bizimhesap_url: Optional[str] = None
        bizimhesap_status: Optional[str] = None
        bizimhesap_error: Optional[str] = None

    class TransactionCreate(BaseModel):
        type: str
        category: str
        amount: float
        description: str
        reference_id: Optional[str] = ""
        reference_type: Optional[str] = ""

    class TransactionResponse(BaseModel):
        id: str
        type: str
        category: str
        amount: float
        description: str
        reference_id: str
        reference_type: str
        created_at: str

    class StockMovementCreate(BaseModel):
        material_id: str
        material_name: str
        type: str
        quantity: float
        reason: str
        reference_id: Optional[str] = ""

    class StockMovementResponse(BaseModel):
        id: str
        material_id: str
        material_name: str
        type: str
        quantity: float
        reason: str
        reference_id: str
        created_at: str

    class DealerLogin(BaseModel):
        dealer_code: str
        password: str

    class DealerOrderCreate(BaseModel):
        items: List[OrderItem]
        delivery_date: str
        notes: Optional[str] = ""

    # ==================== AUTH HELPERS ====================

    def hash_password(password: str) -> str:
        return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

    def verify_password(password: str, hashed: str) -> bool:
        return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

    def create_token(user_id: str, email: str, role: str) -> str:
        from datetime import timedelta
        payload = {
            "sub": user_id,
            "email": email,
            "role": role,
            "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
        }
        return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

    async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
        try:
            payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
            if not user:
                raise HTTPException(status_code=401, detail="Kullanıcı bulunamadı")
            return user
        except jwt.ExpiredSignatureError:
            raise HTTPException(status_code=401, detail="Token süresi dolmuş")
        except jwt.InvalidTokenError:
            raise HTTPException(status_code=401, detail="Geçersiz token")

    async def get_current_dealer(credentials: HTTPAuthorizationCredentials = Depends(security)):
        try:
            payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            if payload.get("type") != "dealer":
                raise HTTPException(status_code=401, detail="Bayi yetkisi gerekli")
            dealer = await db.dealers.find_one({"id": payload["sub"]}, {"_id": 0})
            if not dealer:
                raise HTTPException(status_code=401, detail="Bayi bulunamadı")
            return dealer
        except jwt.ExpiredSignatureError:
            raise HTTPException(status_code=401, detail="Token süresi dolmuş")
        except jwt.InvalidTokenError:
            raise HTTPException(status_code=401, detail="Geçersiz token")

    # ==================== AUTH ROUTES ====================

    # Public registration is DISABLED for security
    # Only admins can create new users via /api/admin/users endpoint
    
    @api_router.post("/auth/register", response_model=TokenResponse)
    async def register(user_data: UserCreate):
        """Public registration is disabled. Only admins can create users."""
        raise HTTPException(
            status_code=403, 
            detail="Kayıt devre dışı. Yeni kullanıcı eklemek için admin ile iletişime geçin."
        )

    @api_router.post("/admin/users", response_model=TokenResponse)
    async def admin_create_user(user_data: UserCreate, current_user: dict = Depends(get_current_user)):
        """Only admins can create new users"""
        if current_user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
        
        existing = await db.users.find_one({"email": user_data.email})
        if existing:
            raise HTTPException(status_code=400, detail="Bu email zaten kayıtlı")
        
        user_id = str(uuid.uuid4())
        user_doc = {
            "id": user_id,
            "email": user_data.email,
            "password": hash_password(user_data.password),
            "name": user_data.name,
            "role": user_data.role,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user["id"]
        }
        await db.users.insert_one(user_doc)
        
        token = create_token(user_id, user_data.email, user_data.role)
        user_response = UserResponse(
            id=user_id,
            email=user_data.email,
            name=user_data.name,
            role=user_data.role,
            created_at=user_doc["created_at"]
        )
        return TokenResponse(access_token=token, user=user_response)

    @api_router.get("/admin/users")
    async def admin_list_users(current_user: dict = Depends(get_current_user)):
        """List all users - admin only"""
        if current_user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
        users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(100)
        return users

    @api_router.delete("/admin/users/{user_id}")
    async def admin_delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
        """Delete a user - admin only"""
        if current_user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
        if user_id == current_user["id"]:
            raise HTTPException(status_code=400, detail="Kendinizi silemezsiniz")
        result = await db.users.delete_one({"id": user_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
        return {"message": "Kullanıcı silindi"}

    # ==================== CAPTCHA ENDPOINTS ====================
    
    @api_router.get("/auth/captcha")
    async def get_captcha():
        """Yeni captcha al"""
        captcha = generate_captcha()
        captcha_storage[captcha["id"]] = {
            "answer": captcha["answer"],
            "created_at": time.time()
        }
        # 5 dakika sonra expire olacak captcha'ları temizle
        expired = [k for k, v in captcha_storage.items() if time.time() - v["created_at"] > 300]
        for k in expired:
            del captcha_storage[k]
        return {"captcha_id": captcha["id"], "question": captcha["question"]}

    @api_router.post("/auth/login")
    async def login(credentials: UserLogin, request: Request, captcha_id: Optional[str] = None):
        client_ip = request.client.host if request.client else "unknown"
        current_time = time.time()
        
        # Check if IP is blocked due to too many failed attempts
        if client_ip in blocked_ips:
            if current_time < blocked_ips[client_ip]:
                remaining = int(blocked_ips[client_ip] - current_time)
                raise HTTPException(
                    status_code=429, 
                    detail=f"Çok fazla başarısız giriş denemesi. {remaining} saniye sonra tekrar deneyin."
                )
            else:
                del blocked_ips[client_ip]
                failed_login_attempts[client_ip] = []
        
        # Clean old failed attempts
        failed_login_attempts[client_ip] = [
            t for t in failed_login_attempts[client_ip] 
            if current_time - t < LOGIN_BLOCK_DURATION
        ]
        
        # Captcha kontrolü - 2 veya daha fazla başarısız deneme varsa zorunlu
        if len(failed_login_attempts[client_ip]) >= 2:
            if not captcha_id or not credentials.captcha_answer:
                raise HTTPException(
                    status_code=400, 
                    detail="Captcha doğrulaması gerekli",
                    headers={"X-Captcha-Required": "true"}
                )
            
            if captcha_id not in captcha_storage:
                raise HTTPException(status_code=400, detail="Geçersiz veya süresi dolmuş captcha")
            
            if captcha_storage[captcha_id]["answer"] != credentials.captcha_answer:
                raise HTTPException(status_code=400, detail="Yanlış captcha yanıtı")
            
            # Kullanılmış captcha'yı sil
            del captcha_storage[captcha_id]
        
        user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
        if not user or not verify_password(credentials.password, user["password"]):
            # Record failed attempt
            failed_login_attempts[client_ip].append(current_time)
            
            # Log failed attempt
            log_audit("LOGIN_FAILED", "unknown", credentials.email, client_ip, "Invalid credentials")
            
            # Block IP if too many failed attempts
            if len(failed_login_attempts[client_ip]) >= LOGIN_ATTEMPT_LIMIT:
                blocked_ips[client_ip] = current_time + LOGIN_BLOCK_DURATION
                logging.warning(f"IP blocked due to failed login attempts: {client_ip}")
                log_audit("IP_BLOCKED", "unknown", credentials.email, client_ip, "Too many failed attempts")
                raise HTTPException(
                    status_code=429, 
                    detail=f"Çok fazla başarısız giriş denemesi. {LOGIN_BLOCK_DURATION // 60} dakika engellendi."
                )
            
            remaining_attempts = LOGIN_ATTEMPT_LIMIT - len(failed_login_attempts[client_ip])
            captcha_required = len(failed_login_attempts[client_ip]) >= 2
            raise HTTPException(
                status_code=401, 
                detail=f"Geçersiz email veya şifre. Kalan deneme: {remaining_attempts}",
                headers={"X-Captcha-Required": str(captcha_required).lower()}
            )
        
        # Check if 2FA is enabled for this user
        if user.get("two_factor_enabled", False):
            # Generate and store 2FA code
            code = generate_2fa_code()
            two_factor_codes[user["email"]] = {
                "code": code,
                "created_at": time.time(),
                "user_id": user["id"]
            }
            
            # Log 2FA code generation
            log_audit("2FA_CODE_GENERATED", user["id"], user["email"], client_ip)
            logging.info(f"2FA code for {user['email']}: {code}")  # Production'da email ile gönderilecek
            
            return {
                "requires_2fa": True,
                "message": "2FA kodu email adresinize gönderildi",
                "email": user["email"]
            }
        
        # Clear failed attempts on successful login
        failed_login_attempts[client_ip] = []
        
        token = create_token(user["id"], user["email"], user["role"])
        user_response = UserResponse(
            id=user["id"],
            email=user["email"],
            name=user["name"],
            role=user["role"],
            created_at=user["created_at"],
            two_factor_enabled=user.get("two_factor_enabled", False)
        )
        
        # Log successful login
        log_audit("LOGIN_SUCCESS", user["id"], user["email"], client_ip)
        logging.info(f"Successful login: {user['email']} from IP: {client_ip}")
        
        return {"access_token": token, "token_type": "bearer", "user": user_response, "requires_2fa": False}

    @api_router.post("/auth/verify-2fa")
    async def verify_2fa(data: TwoFactorVerify, request: Request):
        """2FA kodunu doğrula"""
        client_ip = request.client.host if request.client else "unknown"
        
        if data.email not in two_factor_codes:
            raise HTTPException(status_code=400, detail="2FA kodu bulunamadı veya süresi dolmuş")
        
        stored = two_factor_codes[data.email]
        
        # 5 dakika içinde girilmeli
        if time.time() - stored["created_at"] > 300:
            del two_factor_codes[data.email]
            raise HTTPException(status_code=400, detail="2FA kodunun süresi dolmuş")
        
        if stored["code"] != data.code:
            log_audit("2FA_FAILED", stored["user_id"], data.email, client_ip, "Invalid code")
            raise HTTPException(status_code=400, detail="Geçersiz 2FA kodu")
        
        # Kod doğru - token oluştur
        del two_factor_codes[data.email]
        
        user = await db.users.find_one({"email": data.email}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
        
        token = create_token(user["id"], user["email"], user["role"])
        user_response = UserResponse(
            id=user["id"],
            email=user["email"],
            name=user["name"],
            role=user["role"],
            created_at=user["created_at"],
            two_factor_enabled=True
        )
        
        log_audit("2FA_SUCCESS", user["id"], data.email, client_ip)
        
        return {"access_token": token, "token_type": "bearer", "user": user_response}

    @api_router.post("/auth/toggle-2fa")
    async def toggle_2fa(current_user: dict = Depends(get_current_user)):
        """2FA'yı aç/kapat"""
        user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0})
        current_status = user.get("two_factor_enabled", False)
        new_status = not current_status
        
        await db.users.update_one(
            {"id": current_user["id"]},
            {"$set": {"two_factor_enabled": new_status}}
        )
        
        log_audit("2FA_TOGGLED", current_user["id"], current_user["email"], "system", f"2FA {'enabled' if new_status else 'disabled'}")
        
        return {
            "two_factor_enabled": new_status,
            "message": f"2FA {'aktif edildi' if new_status else 'devre dışı bırakıldı'}"
        }

    @api_router.get("/auth/me", response_model=UserResponse)
    async def get_me(current_user: dict = Depends(get_current_user)):
        return UserResponse(
            id=current_user["id"],
            email=current_user["email"],
            name=current_user["name"],
            role=current_user["role"],
            created_at=current_user["created_at"]
        )

    class ChangePasswordRequest(BaseModel):
        current_password: str
        new_password: str

    @api_router.put("/auth/change-password")
    async def change_password(request: ChangePasswordRequest, current_user: dict = Depends(get_current_user)):
        """Admin kullanıcısının şifresini değiştir"""
        # Mevcut kullanıcıyı veritabanından al (şifre dahil)
        user = await db.users.find_one({"id": current_user["id"]})
        if not user:
            raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
        
        # Mevcut şifreyi doğrula
        if not verify_password(request.current_password, user["password"]):
            raise HTTPException(status_code=400, detail="Mevcut şifre yanlış")
        
        # Şifre politikası kontrolü
        is_valid, message = validate_password_strength(request.new_password)
        if not is_valid:
            raise HTTPException(status_code=400, detail=message)
        
        # Yeni şifreyi hashle ve güncelle
        hashed_password = hash_password(request.new_password)
        await db.users.update_one(
            {"id": current_user["id"]},
            {"$set": {"password": hashed_password}}
        )
        
        # Audit log
        log_audit("PASSWORD_CHANGED", current_user["id"], current_user["email"], "system")
        
        return {"message": "Şifre başarıyla değiştirildi"}

    # ==================== PRODUCT ROUTES ====================

    @api_router.post("/products", response_model=ProductResponse)
    async def create_product(product: ProductCreate, current_user: dict = Depends(get_current_user)):
        product_id = str(uuid.uuid4())
        product_doc = {
            "id": product_id,
            **product.model_dump(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.products.insert_one(product_doc)
        return ProductResponse(**{k: v for k, v in product_doc.items() if k != "_id"})

    @api_router.get("/products", response_model=List[ProductResponse])
    async def get_products(current_user: dict = Depends(get_current_user)):
        products = await db.products.find({}, {"_id": 0}).to_list(1000)
        return [ProductResponse(**p) for p in products]

    @api_router.get("/products/{product_id}", response_model=ProductResponse)
    async def get_product(product_id: str, current_user: dict = Depends(get_current_user)):
        product = await db.products.find_one({"id": product_id}, {"_id": 0})
        if not product:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        return ProductResponse(**product)

    @api_router.put("/products/{product_id}", response_model=ProductResponse)
    async def update_product(product_id: str, product: ProductCreate, current_user: dict = Depends(get_current_user)):
        result = await db.products.update_one({"id": product_id}, {"$set": product.model_dump()})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        updated = await db.products.find_one({"id": product_id}, {"_id": 0})
        return ProductResponse(**updated)

    @api_router.delete("/products/{product_id}")
    async def delete_product(product_id: str, current_user: dict = Depends(get_current_user)):
        result = await db.products.delete_one({"id": product_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        return {"message": "Ürün silindi"}

    # ==================== MATERIAL ROUTES ====================

    @api_router.post("/materials", response_model=MaterialResponse)
    async def create_material(material: MaterialCreate, current_user: dict = Depends(get_current_user)):
        material_id = str(uuid.uuid4())
        material_doc = {
            "id": material_id,
            **material.model_dump(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.materials.insert_one(material_doc)
        return MaterialResponse(**{k: v for k, v in material_doc.items() if k != "_id"})

    @api_router.get("/materials", response_model=List[MaterialResponse])
    async def get_materials(current_user: dict = Depends(get_current_user)):
        materials = await db.materials.find({}, {"_id": 0}).to_list(1000)
        return [MaterialResponse(**m) for m in materials]

    @api_router.put("/materials/{material_id}", response_model=MaterialResponse)
    async def update_material(material_id: str, material: MaterialCreate, current_user: dict = Depends(get_current_user)):
        result = await db.materials.update_one({"id": material_id}, {"$set": material.model_dump()})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Hammadde bulunamadı")
        updated = await db.materials.find_one({"id": material_id}, {"_id": 0})
        return MaterialResponse(**updated)

    @api_router.delete("/materials/{material_id}")
    async def delete_material(material_id: str, current_user: dict = Depends(get_current_user)):
        result = await db.materials.delete_one({"id": material_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Hammadde bulunamadı")
        return {"message": "Hammadde silindi"}

    # ==================== RECIPE ROUTES ====================

    @api_router.post("/recipes", response_model=RecipeResponse)
    async def create_recipe(recipe: RecipeCreate, current_user: dict = Depends(get_current_user)):
        recipe_id = str(uuid.uuid4())
        recipe_doc = {
            "id": recipe_id,
            **recipe.model_dump(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.recipes.insert_one(recipe_doc)
        return RecipeResponse(**{k: v for k, v in recipe_doc.items() if k != "_id"})

    SUPER_ADMIN_EMAIL = "admin@kasaburger.net.tr"
    
    @api_router.get("/recipes", response_model=List[dict])
    async def get_recipes(current_user: dict = Depends(get_current_user)):
        # Only super admin can access recipes
        if current_user.get("email") != SUPER_ADMIN_EMAIL:
            raise HTTPException(status_code=403, detail="Bu sayfaya erişim yetkiniz yok")
        recipes = await db.recipes.find({}, {"_id": 0}).to_list(1000)
        return recipes

    @api_router.delete("/recipes/{recipe_id}")
    async def delete_recipe(recipe_id: str, current_user: dict = Depends(get_current_user)):
        result = await db.recipes.delete_one({"id": recipe_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Reçete bulunamadı")
        return {"message": "Reçete silindi"}

    # ==================== PRODUCTION ROUTES ====================

    @api_router.post("/production", response_model=ProductionResponse)
    async def create_production(production: ProductionCreate, current_user: dict = Depends(get_current_user)):
        production_id = str(uuid.uuid4())
        production_doc = {
            "id": production_id,
            **production.model_dump(),
            "status": "planned",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "completed_at": None
        }
        await db.production.insert_one(production_doc)
        return ProductionResponse(**{k: v for k, v in production_doc.items() if k != "_id"})

    @api_router.get("/production", response_model=List[ProductionResponse])
    async def get_productions(current_user: dict = Depends(get_current_user)):
        productions = await db.production.find({}, {"_id": 0}).to_list(1000)
        return [ProductionResponse(**p) for p in productions]

    @api_router.put("/production/{production_id}/status")
    async def update_production_status(production_id: str, status: str, current_user: dict = Depends(get_current_user)):
        update_data = {"status": status}
        if status == "completed":
            update_data["completed_at"] = datetime.now(timezone.utc).isoformat()
        result = await db.production.update_one({"id": production_id}, {"$set": update_data})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Üretim emri bulunamadı")
        return {"message": "Durum güncellendi"}

    @api_router.delete("/production/{production_id}")
    async def delete_production(production_id: str, current_user: dict = Depends(get_current_user)):
        result = await db.production.delete_one({"id": production_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Üretim emri bulunamadı")
        return {"message": "Üretim emri silindi"}

    # ==================== DEALER ROUTES ====================

    @api_router.post("/dealers", response_model=DealerResponse)
    async def create_dealer(dealer: DealerCreate, current_user: dict = Depends(get_current_user)):
        dealer_id = str(uuid.uuid4())
        dealer_data = dealer.model_dump()
        # Şifre belirtilmemişse bayi kodunu kullan
        if not dealer_data.get("password"):
            dealer_data["password"] = dealer_data["code"]
        dealer_doc = {
            "id": dealer_id,
            **dealer_data,
            "balance": 0,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.dealers.insert_one(dealer_doc)
        return DealerResponse(**{k: v for k, v in dealer_doc.items() if k != "_id" and k != "password"})

    @api_router.get("/dealers", response_model=List[DealerResponse])
    async def get_dealers(current_user: dict = Depends(get_current_user)):
        dealers = await db.dealers.find({}, {"_id": 0}).to_list(1000)
        return [DealerResponse(**d) for d in dealers]

    @api_router.get("/dealers/{dealer_id}", response_model=DealerResponse)
    async def get_dealer(dealer_id: str, current_user: dict = Depends(get_current_user)):
        dealer = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
        if not dealer:
            raise HTTPException(status_code=404, detail="Bayi bulunamadı")
        return DealerResponse(**dealer)

    @api_router.put("/dealers/{dealer_id}", response_model=DealerResponse)
    async def update_dealer(dealer_id: str, dealer: DealerCreate, current_user: dict = Depends(get_current_user)):
        update_data = dealer.model_dump()
        # Şifre boşsa güncelleme verisinden çıkar (mevcut şifreyi koru)
        if not update_data.get("password"):
            update_data.pop("password", None)
        result = await db.dealers.update_one({"id": dealer_id}, {"$set": update_data})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Bayi bulunamadı")
        updated = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
        return DealerResponse(**{k: v for k, v in updated.items() if k != "password"})

    @api_router.delete("/dealers/{dealer_id}")
    async def delete_dealer(dealer_id: str, current_user: dict = Depends(get_current_user)):
        result = await db.dealers.delete_one({"id": dealer_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Bayi bulunamadı")
        return {"message": "Bayi silindi"}

    # ==================== ORDER ROUTES ====================

    async def generate_order_number():
        count = await db.orders.count_documents({})
        return f"SIP-{str(count + 1).zfill(6)}"

    @api_router.post("/orders", response_model=OrderResponse)
    async def create_order(order: OrderCreate, current_user: dict = Depends(get_current_user)):
        order_id = str(uuid.uuid4())
        order_number = await generate_order_number()
        subtotal = sum(item.total for item in order.items)
        tax_amount = subtotal * 0.20
        total = subtotal + tax_amount
        order_doc = {
            "id": order_id,
            "order_number": order_number,
            **order.model_dump(),
            "subtotal": subtotal,
            "tax_amount": tax_amount,
            "total": total,
            "status": "pending",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.orders.insert_one(order_doc)
        return OrderResponse(**{k: v for k, v in order_doc.items() if k != "_id"})

    @api_router.get("/orders", response_model=List[OrderResponse])
    async def get_orders(current_user: dict = Depends(get_current_user)):
        orders = await db.orders.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
        return [OrderResponse(**o) for o in orders]

    @api_router.put("/orders/{order_id}/status")
    async def update_order_status(order_id: str, status: str, current_user: dict = Depends(get_current_user)):
        result = await db.orders.update_one({"id": order_id}, {"$set": {"status": status}})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
        
        # Sipariş onaylandığında otomatik fatura oluştur ve Bizim Hesap'a gönder
        if status == "confirmed":
            order = await db.orders.find_one({"id": order_id}, {"_id": 0})
            if order:
                try:
                    # Fatura oluştur
                    invoice_result = await create_invoice_from_order(order, current_user)
                    
                    # Bizim Hesap'a gönder
                    bizimhesap_result = await send_to_bizimhesap(invoice_result, order)
                    
                    return {
                        "message": "Sipariş onaylandı, fatura oluşturuldu",
                        "invoice_id": invoice_result.get("id"),
                        "invoice_number": invoice_result.get("invoice_number"),
                        "bizimhesap_status": bizimhesap_result.get("status", "pending")
                    }
                except Exception as e:
                    logging.error(f"Fatura oluşturma hatası: {str(e)}")
                    return {"message": "Sipariş onaylandı, fatura oluşturulamadı", "error": str(e)}
        
        return {"message": "Durum güncellendi"}

    async def create_invoice_from_order(order: dict, current_user: dict):
        """Siparişten otomatik fatura oluşturur"""
        invoice_id = str(uuid.uuid4())
        invoice_number = await generate_invoice_number()
        
        # Bayi bilgilerini al
        dealer = await db.dealers.find_one({"id": order["dealer_id"]}, {"_id": 0})
        
        # Vade tarihi (30 gün sonra)
        now = datetime.now(timezone.utc)
        due_date = (now + timedelta(days=30)).strftime("%Y-%m-%d")
        
        invoice_doc = {
            "id": invoice_id,
            "invoice_number": invoice_number,
            "order_id": order["id"],
            "order_number": order.get("order_number", ""),
            "dealer_id": order["dealer_id"],
            "dealer_name": order["dealer_name"],
            "items": order["items"],
            "subtotal": order["subtotal"],
            "tax_rate": 20.0,
            "tax_amount": order["tax_amount"],
            "total": order["total"],
            "due_date": due_date,
            "status": "unpaid",
            "bizimhesap_guid": None,
            "bizimhesap_url": None,
            "bizimhesap_status": None,
            "bizimhesap_error": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user["name"],
            "paid_at": None
        }
        
        await db.invoices.insert_one(invoice_doc)
        await db.dealers.update_one({"id": order["dealer_id"]}, {"$inc": {"balance": order["total"]}})
        
        return {k: v for k, v in invoice_doc.items() if k != "_id"}

    async def send_to_bizimhesap(invoice: dict, order: dict):
        """Faturayı Bizim Hesap'a gönderir"""
        if not BIZIMHESAP_API_KEY:
            return {"status": "skipped", "message": "API key not configured"}
        
        try:
            # Bayi bilgilerini al
            dealer = await db.dealers.find_one({"id": invoice["dealer_id"]}, {"_id": 0})
            
            # Şirket ayarlarını al
            settings = await db.settings.find_one({}, {"_id": 0})
            
            # Bizim Hesap API formatına dönüştür
            now = datetime.now(timezone.utc)
            due_date = now + timedelta(days=30)
            
            # Ürün detayları
            details = []
            for item in invoice["items"]:
                unit_price = item["unit_price"]
                quantity = item["quantity"]
                gross = unit_price * quantity
                tax_rate = 20.0  # KDV oranı
                net = gross
                tax = net * (tax_rate / 100)
                total = net + tax
                
                details.append({
                    "productId": item.get("product_id", ""),
                    "productName": item["product_name"],
                    "note": "",
                    "barcode": "",
                    "taxRate": f"{tax_rate:.2f}",
                    "quantity": quantity,
                    "unitPrice": f"{unit_price:.2f}",
                    "grossPrice": f"{gross:.2f}",
                    "discount": "0.00",
                    "net": f"{net:.2f}",
                    "tax": f"{tax:.2f}",
                    "total": f"{total:.2f}"
                })
            
            payload = {
                "firmId": BIZIMHESAP_API_KEY,
                "invoiceNo": invoice["invoice_number"],
                "invoiceType": 3,  # 3 = Satış faturası
                "note": f"Sipariş No: {order.get('order_number', '')}",
                "dates": {
                    "invoiceDate": now.strftime("%Y-%m-%dT%H:%M:%S.000+03:00"),
                    "dueDate": due_date.strftime("%Y-%m-%dT%H:%M:%S.000+03:00"),
                    "deliveryDate": now.strftime("%Y-%m-%dT%H:%M:%S.000+03:00")
                },
                "customer": {
                    "customerId": dealer.get("id", "") if dealer else "",
                    "title": dealer.get("name", "") if dealer else invoice.get("dealer_name", ""),
                    "taxOffice": dealer.get("tax_office", "") if dealer else "",
                    "taxNo": dealer.get("tax_number", "") if dealer else "",
                    "email": dealer.get("email", "") if dealer else "",
                    "phone": dealer.get("phone", "") if dealer else "",
                    "address": dealer.get("address", "") if dealer else ""
                },
                "amounts": {
                    "currency": "TL",
                    "gross": f"{invoice['subtotal']:.2f}",
                    "discount": "0.00",
                    "net": f"{invoice['subtotal']:.2f}",
                    "tax": f"{invoice['tax_amount']:.2f}",
                    "total": f"{invoice['total']:.2f}"
                },
                "details": details
            }
            
            async with httpx.AsyncClient(timeout=30.0) as client:
                response = await client.post(
                    f"{BIZIMHESAP_API_URL}/addinvoice",
                    json=payload,
                    headers={"Content-Type": "application/json"}
                )
                
                result = response.json()
                
                if result.get("guid"):
                    # Başarılı - faturayı güncelle
                    await db.invoices.update_one(
                        {"id": invoice["id"]},
                        {"$set": {
                            "bizimhesap_guid": result["guid"],
                            "bizimhesap_url": result.get("url", ""),
                            "bizimhesap_status": "sent"
                        }}
                    )
                    return {"status": "success", "guid": result["guid"], "url": result.get("url", "")}
                else:
                    # Hata
                    await db.invoices.update_one(
                        {"id": invoice["id"]},
                        {"$set": {"bizimhesap_status": "error", "bizimhesap_error": result.get("error", "Bilinmeyen hata")}}
                    )
                    return {"status": "error", "message": result.get("error", "Bilinmeyen hata")}
                    
        except Exception as e:
            logging.error(f"Bizim Hesap API hatası: {str(e)}")
            await db.invoices.update_one(
                {"id": invoice["id"]},
                {"$set": {"bizimhesap_status": "error", "bizimhesap_error": str(e)}}
            )
            return {"status": "error", "message": str(e)}

    @api_router.post("/invoices/{invoice_id}/send-bizimhesap")
    async def resend_to_bizimhesap(invoice_id: str, current_user: dict = Depends(get_current_user)):
        """Faturayı tekrar Bizim Hesap'a gönderir"""
        invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
        if not invoice:
            raise HTTPException(status_code=404, detail="Fatura bulunamadı")
        
        order = await db.orders.find_one({"id": invoice.get("order_id")}, {"_id": 0})
        if not order:
            order = {"order_number": invoice.get("invoice_number", "")}
        
        result = await send_to_bizimhesap(invoice, order)
        return result

    @api_router.delete("/orders/{order_id}")
    async def delete_order(order_id: str, current_user: dict = Depends(get_current_user)):
        result = await db.orders.delete_one({"id": order_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
        return {"message": "Sipariş silindi"}

    # ==================== INVOICE ROUTES ====================

    async def generate_invoice_number():
        count = await db.invoices.count_documents({})
        return f"FTR-{str(count + 1).zfill(6)}"

    @api_router.post("/invoices", response_model=InvoiceResponse)
    async def create_invoice(invoice: InvoiceCreate, current_user: dict = Depends(get_current_user)):
        invoice_id = str(uuid.uuid4())
        invoice_number = await generate_invoice_number()
        invoice_doc = {
            "id": invoice_id,
            "invoice_number": invoice_number,
            **invoice.model_dump(),
            "status": "unpaid",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "paid_at": None
        }
        await db.invoices.insert_one(invoice_doc)
        await db.dealers.update_one({"id": invoice.dealer_id}, {"$inc": {"balance": invoice.total}})
        return InvoiceResponse(**{k: v for k, v in invoice_doc.items() if k != "_id"})

    @api_router.get("/invoices", response_model=List[InvoiceResponse])
    async def get_invoices(current_user: dict = Depends(get_current_user)):
        invoices = await db.invoices.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
        return [InvoiceResponse(**i) for i in invoices]

    @api_router.put("/invoices/{invoice_id}/pay")
    async def pay_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
        invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
        if not invoice:
            raise HTTPException(status_code=404, detail="Fatura bulunamadı")
        await db.invoices.update_one({"id": invoice_id}, {"$set": {"status": "paid", "paid_at": datetime.now(timezone.utc).isoformat()}})
        await db.dealers.update_one({"id": invoice["dealer_id"]}, {"$inc": {"balance": -invoice["total"]}})
        await db.transactions.insert_one({
            "id": str(uuid.uuid4()),
            "type": "income",
            "category": "Satış",
            "amount": invoice["total"],
            "description": f"Fatura ödemesi: {invoice['invoice_number']}",
            "reference_id": invoice_id,
            "reference_type": "invoice",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        return {"message": "Fatura ödendi"}

    @api_router.delete("/invoices/{invoice_id}")
    async def delete_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
        invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
        if not invoice:
            raise HTTPException(status_code=404, detail="Fatura bulunamadı")
        if invoice["status"] == "unpaid":
            await db.dealers.update_one({"id": invoice["dealer_id"]}, {"$inc": {"balance": -invoice["total"]}})
        await db.invoices.delete_one({"id": invoice_id})
        return {"message": "Fatura silindi"}

    # ==================== TRANSACTION ROUTES ====================

    @api_router.post("/transactions", response_model=TransactionResponse)
    async def create_transaction(transaction: TransactionCreate, current_user: dict = Depends(get_current_user)):
        transaction_id = str(uuid.uuid4())
        transaction_doc = {
            "id": transaction_id,
            **transaction.model_dump(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.transactions.insert_one(transaction_doc)
        return TransactionResponse(**{k: v for k, v in transaction_doc.items() if k != "_id"})

    @api_router.get("/transactions", response_model=List[TransactionResponse])
    async def get_transactions(current_user: dict = Depends(get_current_user)):
        transactions = await db.transactions.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
        return [TransactionResponse(**t) for t in transactions]

    @api_router.delete("/transactions/{transaction_id}")
    async def delete_transaction(transaction_id: str, current_user: dict = Depends(get_current_user)):
        result = await db.transactions.delete_one({"id": transaction_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="İşlem bulunamadı")
        return {"message": "İşlem silindi"}

    # ==================== STOCK MOVEMENT ROUTES ====================

    @api_router.post("/stock-movements", response_model=StockMovementResponse)
    async def create_stock_movement(movement: StockMovementCreate, current_user: dict = Depends(get_current_user)):
        movement_id = str(uuid.uuid4())
        movement_doc = {
            "id": movement_id,
            **movement.model_dump(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.stock_movements.insert_one(movement_doc)
        quantity_change = movement.quantity if movement.type == "in" else -movement.quantity
        await db.materials.update_one({"id": movement.material_id}, {"$inc": {"stock_quantity": quantity_change}})
        return StockMovementResponse(**{k: v for k, v in movement_doc.items() if k != "_id"})

    @api_router.get("/stock-movements", response_model=List[StockMovementResponse])
    async def get_stock_movements(current_user: dict = Depends(get_current_user)):
        movements = await db.stock_movements.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
        return [StockMovementResponse(**m) for m in movements]

    # ==================== DASHBOARD ROUTES ====================

    @api_router.get("/dashboard/stats")
    async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
        total_products = await db.products.count_documents({})
        total_dealers = await db.dealers.count_documents({})
        total_orders = await db.orders.count_documents({})
        pending_orders = await db.orders.count_documents({"status": "pending"})
        invoices = await db.invoices.find({}, {"_id": 0, "total": 1, "status": 1}).to_list(1000)
        total_revenue = sum(i["total"] for i in invoices if i["status"] == "paid")
        unpaid_invoices = sum(i["total"] for i in invoices if i["status"] == "unpaid")
        transactions = await db.transactions.find({}, {"_id": 0, "type": 1, "amount": 1}).to_list(1000)
        total_income = sum(t["amount"] for t in transactions if t["type"] == "income")
        total_expense = sum(t["amount"] for t in transactions if t["type"] == "expense")
        low_stock_materials = await db.materials.find({"$expr": {"$lte": ["$stock_quantity", "$min_stock"]}}, {"_id": 0}).to_list(100)
        recent_orders = await db.orders.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
        planned_production = await db.production.count_documents({"status": "planned"})
        in_progress_production = await db.production.count_documents({"status": "in_progress"})
        return {
            "total_products": total_products,
            "total_dealers": total_dealers,
            "total_orders": total_orders,
            "pending_orders": pending_orders,
            "total_revenue": total_revenue,
            "unpaid_invoices": unpaid_invoices,
            "total_income": total_income,
            "total_expense": total_expense,
            "net_profit": total_income - total_expense,
            "low_stock_materials": low_stock_materials,
            "recent_orders": recent_orders,
            "planned_production": planned_production,
            "in_progress_production": in_progress_production
        }

    # ==================== PDF & EXCEL EXPORT ====================

    @api_router.get("/invoices/{invoice_id}/pdf")
    async def get_invoice_pdf(invoice_id: str, current_user: dict = Depends(get_current_user)):
        if not reportlab_available:
            raise HTTPException(status_code=503, detail="PDF export not available")
        invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
        if not invoice:
            raise HTTPException(status_code=404, detail="Fatura bulunamadı")
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=24, textColor=colors.HexColor('#f97316'))
        elements.append(Paragraph("KASABURGER", title_style))
        elements.append(Paragraph("Burger Koftesi Imalathanesi", styles['Normal']))
        elements.append(Spacer(1, 10*mm))
        elements.append(Paragraph(f"<b>Fatura No:</b> {invoice['invoice_number']}", styles['Normal']))
        elements.append(Paragraph(f"<b>Tarih:</b> {invoice['created_at'][:10]}", styles['Normal']))
        elements.append(Paragraph(f"<b>Bayi:</b> {invoice['dealer_name']}", styles['Normal']))
        elements.append(Spacer(1, 10*mm))
        table_data = [['Urun', 'Miktar', 'Birim Fiyat', 'Toplam']]
        for item in invoice['items']:
            table_data.append([item['product_name'], str(item['quantity']), f"{item['unit_price']:.2f} TL", f"{item['total']:.2f} TL"])
        table = Table(table_data, colWidths=[80*mm, 25*mm, 35*mm, 35*mm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f97316')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#333333')),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 5*mm))
        elements.append(Paragraph(f"<b>GENEL TOPLAM:</b> {invoice['total']:.2f} TL", styles['Heading2']))
        doc.build(elements)
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=fatura_{invoice['invoice_number']}.pdf"})

    @api_router.get("/reports/excel")
    async def export_reports_excel(report_type: str = "all", current_user: dict = Depends(get_current_user)):
        if not openpyxl_available:
            raise HTTPException(status_code=503, detail="Excel export not available")
        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
        if report_type in ["all", "orders"]:
            ws = wb.active if report_type == "orders" else wb.create_sheet("Siparisler")
            ws.title = "Siparisler"
            orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
            headers = ["Siparis No", "Bayi", "Toplam", "Durum"]
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).font = header_font
                ws.cell(row=1, column=col).fill = header_fill
            for order in orders:
                ws.append([order.get('order_number', ''), order.get('dealer_name', ''), order.get('total', 0), order.get('status', '')])
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=kasaburger_rapor.xlsx"})

    @api_router.get("/invoices/{invoice_id}/xml")
    async def get_invoice_xml(invoice_id: str, current_user: dict = Depends(get_current_user)):
        invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
        if not invoice:
            raise HTTPException(status_code=404, detail="Fatura bulunamadı")
        xml_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<Invoice><ID>{invoice['invoice_number']}</ID><Total>{invoice['total']}</Total></Invoice>'''
        return StreamingResponse(io.BytesIO(xml_content.encode('utf-8')), media_type="application/xml", headers={"Content-Disposition": f"attachment; filename=efatura_{invoice['invoice_number']}.xml"})

    # ==================== DEALER PORTAL ====================

    @api_router.post("/dealer-portal/login")
    async def dealer_portal_login(credentials: DealerLogin, request: Request):
        client_ip = request.client.host if request.client else "unknown"
        current_time = time.time()
        dealer_key = f"dealer_{client_ip}"
        
        # Check if IP is blocked due to too many failed attempts
        if dealer_key in blocked_ips:
            if current_time < blocked_ips[dealer_key]:
                remaining = int(blocked_ips[dealer_key] - current_time)
                raise HTTPException(
                    status_code=429, 
                    detail=f"Çok fazla başarısız giriş denemesi. {remaining} saniye sonra tekrar deneyin."
                )
            else:
                del blocked_ips[dealer_key]
                failed_login_attempts[dealer_key] = []
        
        # Clean old failed attempts
        failed_login_attempts[dealer_key] = [
            t for t in failed_login_attempts[dealer_key] 
            if current_time - t < LOGIN_BLOCK_DURATION
        ]
        
        dealer = await db.dealers.find_one({"code": credentials.dealer_code}, {"_id": 0})
        if not dealer:
            failed_login_attempts[dealer_key].append(current_time)
            if len(failed_login_attempts[dealer_key]) >= LOGIN_ATTEMPT_LIMIT:
                blocked_ips[dealer_key] = current_time + LOGIN_BLOCK_DURATION
                logging.warning(f"Dealer IP blocked due to failed login attempts: {client_ip}")
            raise HTTPException(status_code=401, detail="Geçersiz bayi kodu")
        
        stored_password = dealer.get("password", dealer["code"])
        if credentials.password != stored_password:
            failed_login_attempts[dealer_key].append(current_time)
            if len(failed_login_attempts[dealer_key]) >= LOGIN_ATTEMPT_LIMIT:
                blocked_ips[dealer_key] = current_time + LOGIN_BLOCK_DURATION
                logging.warning(f"Dealer IP blocked due to failed login attempts: {client_ip}")
                raise HTTPException(
                    status_code=429, 
                    detail=f"Çok fazla başarısız giriş denemesi. {LOGIN_BLOCK_DURATION // 60} dakika engellendi."
                )
            remaining_attempts = LOGIN_ATTEMPT_LIMIT - len(failed_login_attempts[dealer_key])
            raise HTTPException(status_code=401, detail=f"Geçersiz şifre. Kalan deneme: {remaining_attempts}")
        
        # Clear failed attempts on successful login
        failed_login_attempts[dealer_key] = []
        
        from datetime import timedelta
        payload = {
            "sub": dealer["id"],
            "dealer_code": dealer["code"],
            "dealer_name": dealer["name"],
            "type": "dealer",
            "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
        }
        token = jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
        
        # Log successful login
        logging.info(f"Successful dealer login: {dealer['code']} from IP: {client_ip}")
        
        return {"access_token": token, "token_type": "bearer", "dealer": {"id": dealer["id"], "name": dealer["name"], "code": dealer["code"], "balance": dealer.get("balance", 0)}}

    @api_router.get("/dealer-portal/me")
    async def dealer_portal_me(dealer: dict = Depends(get_current_dealer)):
        return {"id": dealer["id"], "name": dealer["name"], "code": dealer["code"], "balance": dealer.get("balance", 0), "pricing": dealer.get("pricing", [])}

    @api_router.get("/dealer-portal/products")
    async def dealer_portal_products(dealer: dict = Depends(get_current_dealer)):
        products = await db.products.find({}, {"_id": 0}).to_list(1000)
        dealer_pricing = {p["product_id"]: p["special_price"] for p in dealer.get("pricing", [])}
        return [{**p, "dealer_price": dealer_pricing.get(p["id"], p["base_price"])} for p in products]

    @api_router.get("/dealer-portal/orders")
    async def dealer_portal_orders(dealer: dict = Depends(get_current_dealer)):
        return await db.orders.find({"dealer_id": dealer["id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)

    @api_router.post("/dealer-portal/orders")
    async def dealer_portal_create_order(order: DealerOrderCreate, dealer: dict = Depends(get_current_dealer)):
        # Güncel bayi bilgilerini al (balance ve credit_limit için)
        current_dealer = await db.dealers.find_one({"id": dealer["id"]}, {"_id": 0})
        if not current_dealer:
            raise HTTPException(status_code=404, detail="Bayi bulunamadı")
        
        order_id = str(uuid.uuid4())
        order_number = await generate_order_number()
        subtotal = sum(item.total for item in order.items)
        tax_amount = subtotal * 0.20
        total = subtotal + tax_amount
        
        # Kredi limiti kontrolü
        current_balance = current_dealer.get("balance", 0)  # Mevcut borç
        credit_limit = current_dealer.get("credit_limit", 0)  # Kredi limiti
        new_balance = current_balance + total  # Sipariş sonrası yeni borç
        
        # Limit aşımı kontrolü (credit_limit > 0 ise kontrol yap)
        requires_approval = False
        if credit_limit > 0 and new_balance > credit_limit:
            requires_approval = True
        
        order_doc = {
            "id": order_id, 
            "order_number": order_number, 
            "dealer_id": dealer["id"], 
            "dealer_name": dealer["name"],
            "items": [item.model_dump() for item in order.items], 
            "subtotal": subtotal, 
            "tax_amount": tax_amount,
            "total": total, 
            "delivery_date": order.delivery_date, 
            "notes": order.notes or "", 
            "status": "pending_approval" if requires_approval else "pending",
            "requires_approval": requires_approval,
            "approval_reason": f"Kredi limiti aşımı. Limit: {credit_limit:.2f} TL, Mevcut Borç: {current_balance:.2f} TL, Sipariş Tutarı: {total:.2f} TL" if requires_approval else None,
            "created_at": datetime.now(timezone.utc).isoformat(), 
            "source": "dealer_portal"
        }
        await db.orders.insert_one(order_doc)
        
        response = OrderResponse(**{k: v for k, v in order_doc.items() if k != "_id" and k not in ["requires_approval", "approval_reason"]})
        
        if requires_approval:
            return {
                "order": response.model_dump(),
                "warning": "Kredi limitinizi aştığınız için siparişiniz onay bekliyor.",
                "status": "pending_approval",
                "credit_info": {
                    "credit_limit": credit_limit,
                    "current_balance": current_balance,
                    "order_total": total,
                    "new_balance": new_balance,
                    "over_limit": new_balance - credit_limit
                }
            }
        
        return response

    @api_router.get("/dealer-portal/invoices")
    async def dealer_portal_invoices(dealer: dict = Depends(get_current_dealer)):
        return await db.invoices.find({"dealer_id": dealer["id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)

    @api_router.get("/dealer-portal/campaigns")
    async def dealer_portal_campaigns(dealer: dict = Depends(get_current_dealer)):
        """Bayiye özel aktif kampanyaları getir"""
        now = datetime.now(timezone.utc).isoformat()
        
        # Aktif kampanyaları getir (bitiş tarihi geçmemiş)
        campaigns = await db.campaigns.find({
            "end_date": {"$gte": now[:10]}
        }, {"_id": 0}).sort("created_at", -1).to_list(50)
        
        # Bayiye özel kampanyaları filtrele
        dealer_campaigns = []
        for campaign in campaigns:
            target_dealers = campaign.get("target_dealers", [])
            # Eğer hedef bayi listesi boşsa tüm bayilere, değilse sadece listedeki bayilere
            if len(target_dealers) == 0 or dealer["id"] in target_dealers:
                dealer_campaigns.append(campaign)
        
        return dealer_campaigns

    @api_router.get("/dealer-portal/invoices/{invoice_id}/pdf")
    async def dealer_portal_invoice_pdf(invoice_id: str, dealer: dict = Depends(get_current_dealer)):
        if not reportlab_available:
            raise HTTPException(status_code=503, detail="PDF export not available")
        invoice = await db.invoices.find_one({"id": invoice_id, "dealer_id": dealer["id"]}, {"_id": 0})
        if not invoice:
            raise HTTPException(status_code=404, detail="Fatura bulunamadı")
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=24, textColor=colors.HexColor('#f97316'))
        elements.append(Paragraph("KASABURGER", title_style))
        elements.append(Paragraph(f"<b>Fatura No:</b> {invoice['invoice_number']}", styles['Normal']))
        elements.append(Paragraph(f"<b>Tarih:</b> {invoice['created_at'][:10]}", styles['Normal']))
        elements.append(Spacer(1, 10*mm))
        table_data = [['Urun', 'Miktar', 'Birim Fiyat', 'Toplam']]
        for item in invoice['items']:
            table_data.append([item['product_name'], str(item['quantity']), f"{item['unit_price']:.2f} TL", f"{item['total']:.2f} TL"])
        table = Table(table_data, colWidths=[80*mm, 25*mm, 35*mm, 35*mm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f97316')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#333333')),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 5*mm))
        elements.append(Paragraph(f"<b>GENEL TOPLAM:</b> {invoice['total']:.2f} TL", styles['Heading2']))
        doc.build(elements)
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=fatura_{invoice['invoice_number']}.pdf"})

    @api_router.put("/dealer-portal/change-password")
    async def dealer_change_password(old_password: str, new_password: str, dealer: dict = Depends(get_current_dealer)):
        db_dealer = await db.dealers.find_one({"id": dealer["id"]}, {"_id": 0})
        stored_password = db_dealer.get("password", db_dealer["code"])
        if old_password != stored_password:
            raise HTTPException(status_code=400, detail="Mevcut şifre yanlış")
        await db.dealers.update_one({"id": dealer["id"]}, {"$set": {"password": new_password}})
        return {"message": "Şifre başarıyla değiştirildi"}

    @api_router.get("/dealer-portal/payments")
    async def dealer_portal_payments(dealer: dict = Depends(get_current_dealer)):
        """Bayinin ödeme geçmişini getirir"""
        payments = await db.payments.find({"dealer_id": dealer["id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
        # Fatura numaralarını batch olarak al (N+1 sorgu optimizasyonu)
        invoice_ids = [p["invoice_id"] for p in payments if p.get("invoice_id")]
        invoices = await db.invoices.find({"id": {"$in": invoice_ids}}, {"_id": 0, "id": 1, "invoice_number": 1}).to_list(100)
        invoice_map = {inv["id"]: inv["invoice_number"] for inv in invoices}
        for payment in payments:
            payment["invoice_number"] = invoice_map.get(payment.get("invoice_id"), "-")
        return payments

    class DealerPaymentSubmit(BaseModel):
        amount: float
        payment_method: str = "mail_order"
        payment_date: str
        reference_no: Optional[str] = ""
        notes: Optional[str] = ""

    @api_router.post("/dealer-portal/submit-payment")
    async def dealer_submit_payment(payment: DealerPaymentSubmit, dealer: dict = Depends(get_current_dealer)):
        """Bayi ödeme bildirimi gönderir - Admin onayı bekler"""
        payment_id = str(uuid.uuid4())
        payment_doc = {
            "id": payment_id,
            "dealer_id": dealer["id"],
            "dealer_name": dealer["name"],
            "amount": payment.amount,
            "payment_method": payment.payment_method,
            "payment_date": payment.payment_date,
            "reference_no": payment.reference_no,
            "notes": payment.notes,
            "status": "pending",  # Admin onayı bekliyor
            "submitted_by": "dealer",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.payment_submissions.insert_one(payment_doc)
        return {"message": "Ödeme bildiriminiz alındı", "payment_id": payment_id}

    @api_router.get("/payment-submissions")
    async def get_payment_submissions(current_user: dict = Depends(get_current_user)):
        """Admin için bekleyen ödeme bildirimleri"""
        submissions = await db.payment_submissions.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
        return submissions

    @api_router.put("/payment-submissions/{submission_id}/approve")
    async def approve_payment_submission(submission_id: str, current_user: dict = Depends(get_current_user)):
        """Ödeme bildirimini onayla ve gerçek ödeme olarak kaydet"""
        submission = await db.payment_submissions.find_one({"id": submission_id}, {"_id": 0})
        if not submission:
            raise HTTPException(status_code=404, detail="Ödeme bildirimi bulunamadı")
        if submission["status"] != "pending":
            raise HTTPException(status_code=400, detail="Bu bildirim zaten işlenmiş")
        
        # Gerçek ödeme kaydı oluştur
        payment_id = str(uuid.uuid4())
        payment_doc = {
            "id": payment_id,
            "dealer_id": submission["dealer_id"],
            "dealer_name": submission["dealer_name"],
            "amount": submission["amount"],
            "payment_method": submission["payment_method"],
            "payment_date": submission["payment_date"],
            "reference_no": submission.get("reference_no", ""),
            "notes": f"Bayi bildirimi: {submission.get('notes', '')}",
            "invoice_id": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user["name"]
        }
        await db.payments.insert_one(payment_doc)
        
        # Bayi bakiyesini güncelle
        await db.dealers.update_one({"id": submission["dealer_id"]}, {"$inc": {"balance": -submission["amount"]}})
        
        # Bildirimi onayla
        await db.payment_submissions.update_one({"id": submission_id}, {"$set": {"status": "approved", "approved_by": current_user["name"], "approved_at": datetime.now(timezone.utc).isoformat()}})
        
        return {"message": "Ödeme onaylandı", "payment_id": payment_id}

    @api_router.put("/payment-submissions/{submission_id}/reject")
    async def reject_payment_submission(submission_id: str, reason: str = "", current_user: dict = Depends(get_current_user)):
        """Ödeme bildirimini reddet"""
        result = await db.payment_submissions.update_one(
            {"id": submission_id, "status": "pending"},
            {"$set": {"status": "rejected", "rejected_by": current_user["name"], "rejection_reason": reason, "rejected_at": datetime.now(timezone.utc).isoformat()}}
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Ödeme bildirimi bulunamadı veya zaten işlenmiş")
        return {"message": "Ödeme reddedildi"}

    # ==================== IYZICO SANAL POS ====================
    
    class IyzicoPaymentRequest(BaseModel):
        amount: float
        card_holder_name: str
        card_number: str
        expire_month: str
        expire_year: str
        cvc: str
        installment: int = 1

    @api_router.post("/dealer-portal/iyzico-payment")
    async def process_iyzico_payment(payment: IyzicoPaymentRequest, request: Request, dealer: dict = Depends(get_current_dealer)):
        """iyzico ile kredi kartı ödemesi işle"""
        if not iyzico_available:
            raise HTTPException(status_code=503, detail="Ödeme sistemi şu anda kullanılamıyor")
        
        if not IYZICO_API_KEY or not IYZICO_SECRET_KEY:
            raise HTTPException(status_code=503, detail="Ödeme sistemi yapılandırılmamış")
        
        try:
            # iyzico options
            options = {
                'api_key': IYZICO_API_KEY,
                'secret_key': IYZICO_SECRET_KEY,
                'base_url': IYZICO_BASE_URL
            }
            
            # Generate unique IDs
            conversation_id = str(uuid.uuid4())
            basket_id = f"DEALER-{dealer['id']}-{int(time.time())}"
            
            # Get client IP
            client_ip = request.client.host if request.client else "0.0.0.0"
            
            # Prepare payment request
            payment_request = {
                'locale': 'tr',
                'conversationId': conversation_id,
                'price': str(payment.amount),
                'paidPrice': str(payment.amount),
                'installment': str(payment.installment),
                'basketId': basket_id,
                'paymentChannel': 'WEB',
                'paymentGroup': 'PRODUCT',
                'currency': 'TRY',
                'paymentCard': {
                    'cardHolderName': payment.card_holder_name,
                    'cardNumber': payment.card_number.replace(' ', ''),
                    'expireMonth': payment.expire_month,
                    'expireYear': payment.expire_year,
                    'cvc': payment.cvc,
                    'registerCard': '0'
                },
                'buyer': {
                    'id': dealer['id'],
                    'name': dealer['name'].split()[0] if ' ' in dealer['name'] else dealer['name'],
                    'surname': dealer['name'].split()[-1] if ' ' in dealer['name'] else dealer['name'],
                    'gsmNumber': dealer.get('phone', '+905000000000'),
                    'email': dealer.get('email', 'bayi@kasaburger.com.tr'),
                    'identityNumber': '11111111111',
                    'registrationAddress': dealer.get('address', 'İstanbul'),
                    'ip': client_ip,
                    'city': 'Istanbul',
                    'country': 'Turkey'
                },
                'shippingAddress': {
                    'contactName': dealer['name'],
                    'city': 'Istanbul',
                    'country': 'Turkey',
                    'address': dealer.get('address', 'İstanbul')
                },
                'billingAddress': {
                    'contactName': dealer['name'],
                    'city': 'Istanbul',
                    'country': 'Turkey',
                    'address': dealer.get('address', 'İstanbul')
                },
                'basketItems': [
                    {
                        'id': 'ODEME',
                        'name': 'Bayi Borç Ödemesi',
                        'category1': 'Ödeme',
                        'itemType': 'VIRTUAL',
                        'price': str(payment.amount)
                    }
                ]
            }
            
            # Process payment
            payment_response = iyzipay.Payment().create(payment_request, options)
            response_dict = payment_response.read().decode('utf-8')
            import json
            result = json.loads(response_dict)
            
            if result.get('status') == 'success':
                # Save successful payment
                payment_id = str(uuid.uuid4())
                payment_doc = {
                    "id": payment_id,
                    "dealer_id": dealer["id"],
                    "dealer_name": dealer["name"],
                    "amount": payment.amount,
                    "payment_method": "sanal_pos",
                    "payment_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                    "reference_no": result.get('paymentId', ''),
                    "notes": f"iyzico ödeme - Conversation: {conversation_id}",
                    "iyzico_payment_id": result.get('paymentId'),
                    "card_last_four": payment.card_number[-4:],
                    "invoice_id": None,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": "dealer_portal"
                }
                await db.payments.insert_one(payment_doc)
                
                # Update dealer balance
                await db.dealers.update_one(
                    {"id": dealer["id"]},
                    {"$inc": {"balance": -payment.amount}}
                )
                
                return {
                    "status": "success",
                    "message": "Ödeme başarıyla tamamlandı!",
                    "payment_id": payment_id,
                    "iyzico_payment_id": result.get('paymentId'),
                    "amount": payment.amount
                }
            else:
                # Payment failed
                error_message = result.get('errorMessage', 'Ödeme işlemi başarısız')
                logging.error(f"iyzico payment failed: {result}")
                raise HTTPException(status_code=400, detail=error_message)
                
        except HTTPException:
            raise
        except Exception as e:
            logging.error(f"iyzico payment error: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Ödeme işlemi sırasında hata oluştu: {str(e)}")

    @api_router.post("/dealer-portal/iyzico-bin-check")
    async def check_card_bin(bin_number: str, price: float, dealer: dict = Depends(get_current_dealer)):
        """Kart BIN kontrolü - taksit seçeneklerini getir"""
        if not iyzico_available:
            raise HTTPException(status_code=503, detail="Ödeme sistemi şu anda kullanılamıyor")
        
        try:
            options = {
                'api_key': IYZICO_API_KEY,
                'secret_key': IYZICO_SECRET_KEY,
                'base_url': IYZICO_BASE_URL
            }
            
            request_data = {
                'locale': 'tr',
                'conversationId': str(uuid.uuid4()),
                'binNumber': bin_number[:6],
                'price': str(price)
            }
            
            bin_check = iyzipay.BinNumber().retrieve(request_data, options)
            response_dict = bin_check.read().decode('utf-8')
            import json
            result = json.loads(response_dict)
            
            return result
        except Exception as e:
            logging.error(f"BIN check error: {str(e)}")
            raise HTTPException(status_code=500, detail=str(e))

    # ==================== DEPO/STOK YÖNETİMİ ====================

    class WarehouseCreate(BaseModel):
        name: str
        code: str
        address: Optional[str] = ""
        is_active: bool = True

    @api_router.post("/warehouses")
    async def create_warehouse(warehouse: WarehouseCreate, current_user: dict = Depends(get_current_user)):
        warehouse_id = str(uuid.uuid4())
        warehouse_doc = {
            "id": warehouse_id,
            **warehouse.model_dump(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.warehouses.insert_one(warehouse_doc)
        return {k: v for k, v in warehouse_doc.items() if k != "_id"}

    @api_router.get("/warehouses")
    async def get_warehouses(current_user: dict = Depends(get_current_user)):
        warehouses = await db.warehouses.find({}, {"_id": 0}).to_list(100)
        return warehouses

    @api_router.delete("/warehouses/{warehouse_id}")
    async def delete_warehouse(warehouse_id: str, current_user: dict = Depends(get_current_user)):
        result = await db.warehouses.delete_one({"id": warehouse_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Depo bulunamadı")
        return {"message": "Depo silindi"}

    class StockCountCreate(BaseModel):
        material_id: str
        material_name: str
        counted_quantity: float
        system_quantity: float
        difference: float
        notes: Optional[str] = ""

    @api_router.post("/stock-counts")
    async def create_stock_count(count: StockCountCreate, current_user: dict = Depends(get_current_user)):
        count_id = str(uuid.uuid4())
        count_doc = {
            "id": count_id,
            **count.model_dump(),
            "counted_by": current_user["name"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.stock_counts.insert_one(count_doc)
        if count.difference != 0:
            adjustment_type = "in" if count.difference > 0 else "out"
            await db.stock_movements.insert_one({
                "id": str(uuid.uuid4()),
                "material_id": count.material_id,
                "material_name": count.material_name,
                "type": adjustment_type,
                "quantity": abs(count.difference),
                "reason": f"Stok sayım düzeltmesi: {count.notes}",
                "reference_id": count_id,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            await db.materials.update_one({"id": count.material_id}, {"$set": {"stock_quantity": count.counted_quantity}})
        return {k: v for k, v in count_doc.items() if k != "_id"}

    @api_router.get("/stock-counts")
    async def get_stock_counts(current_user: dict = Depends(get_current_user)):
        counts = await db.stock_counts.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
        return counts

    @api_router.get("/low-stock-alerts")
    async def get_low_stock_alerts(current_user: dict = Depends(get_current_user)):
        alerts = await db.materials.find({"$expr": {"$lte": ["$stock_quantity", "$min_stock"]}}, {"_id": 0}).to_list(100)
        return alerts

    # ==================== BAYİ ŞİFRE YÖNETİMİ ====================

    @api_router.put("/dealers/{dealer_id}/reset-password")
    async def reset_dealer_password(dealer_id: str, new_password: str, current_user: dict = Depends(get_current_user)):
        result = await db.dealers.update_one({"id": dealer_id}, {"$set": {"password": new_password}})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Bayi bulunamadı")
        return {"message": "Bayi şifresi sıfırlandı"}

    # ==================== GELİŞMİŞ RAPORLAR ====================

    @api_router.get("/reports/sales-by-dealer")
    async def sales_by_dealer_report(start_date: Optional[str] = None, end_date: Optional[str] = None, current_user: dict = Depends(get_current_user)):
        query = {}
        if start_date and end_date:
            query["created_at"] = {"$gte": start_date, "$lte": end_date}
        # Sadece gerekli alanları al - performans optimizasyonu
        orders = await db.orders.find(query, {"_id": 0, "dealer_id": 1, "dealer_name": 1, "total": 1}).to_list(10000)
        dealer_sales = {}
        for order in orders:
            dealer_id = order.get("dealer_id", "unknown")
            dealer_name = order.get("dealer_name", "Bilinmeyen")
            if dealer_id not in dealer_sales:
                dealer_sales[dealer_id] = {"dealer_name": dealer_name, "total_orders": 0, "total_amount": 0}
            dealer_sales[dealer_id]["total_orders"] += 1
            dealer_sales[dealer_id]["total_amount"] += order.get("total", 0)
        return list(dealer_sales.values())

    @api_router.get("/reports/sales-by-product")
    async def sales_by_product_report(start_date: Optional[str] = None, end_date: Optional[str] = None, current_user: dict = Depends(get_current_user)):
        query = {}
        if start_date and end_date:
            query["created_at"] = {"$gte": start_date, "$lte": end_date}
        # Sadece gerekli alanları al - performans optimizasyonu
        orders = await db.orders.find(query, {"_id": 0, "items": 1}).to_list(10000)
        product_sales = {}
        for order in orders:
            for item in order.get("items", []):
                product_id = item.get("product_id", "unknown")
                product_name = item.get("product_name", "Bilinmeyen")
                if product_id not in product_sales:
                    product_sales[product_id] = {"product_name": product_name, "total_quantity": 0, "total_amount": 0}
                product_sales[product_id]["total_quantity"] += item.get("quantity", 0)
                product_sales[product_id]["total_amount"] += item.get("total", 0)
        return list(product_sales.values())

    @api_router.get("/reports/monthly-summary")
    async def monthly_summary_report(year: int, month: int, current_user: dict = Depends(get_current_user)):
        start_date = f"{year}-{str(month).zfill(2)}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{str(month + 1).zfill(2)}-01"
        # Sadece gerekli alanları al - performans optimizasyonu
        orders = await db.orders.find({"created_at": {"$gte": start_date, "$lt": end_date}}, {"_id": 0, "total": 1}).to_list(10000)
        invoices = await db.invoices.find({"created_at": {"$gte": start_date, "$lt": end_date}}, {"_id": 0, "total": 1, "status": 1}).to_list(10000)
        transactions = await db.transactions.find({"created_at": {"$gte": start_date, "$lt": end_date}}, {"_id": 0, "type": 1, "amount": 1}).to_list(10000)
        total_orders = len(orders)
        total_order_amount = sum(o.get("total", 0) for o in orders)
        total_invoiced = sum(i.get("total", 0) for i in invoices)
        paid_invoices = sum(i.get("total", 0) for i in invoices if i.get("status") == "paid")
        total_income = sum(t.get("amount", 0) for t in transactions if t.get("type") == "income")
        total_expense = sum(t.get("amount", 0) for t in transactions if t.get("type") == "expense")
        return {
            "year": year,
            "month": month,
            "total_orders": total_orders,
            "total_order_amount": total_order_amount,
            "total_invoiced": total_invoiced,
            "paid_invoices": paid_invoices,
            "unpaid_invoices": total_invoiced - paid_invoices,
            "total_income": total_income,
            "total_expense": total_expense,
            "net_profit": total_income - total_expense
        }

    # ==================== BİLDİRİMLER ====================

    @api_router.get("/notifications")
    async def get_notifications(current_user: dict = Depends(get_current_user)):
        notifications = []
        low_stock = await db.materials.find({"$expr": {"$lte": ["$stock_quantity", "$min_stock"]}}, {"_id": 0}).to_list(100)
        for material in low_stock:
            notifications.append({
                "type": "low_stock",
                "severity": "warning",
                "title": "Düşük Stok Uyarısı",
                "message": f"{material['name']} stoğu kritik seviyede ({material['stock_quantity']} {material['unit']})",
                "material_id": material["id"]
            })
        unpaid_invoices = await db.invoices.find({"status": "unpaid"}, {"_id": 0}).to_list(100)
        for invoice in unpaid_invoices:
            due_date = invoice.get("due_date", "")
            if due_date and due_date < datetime.now(timezone.utc).isoformat()[:10]:
                notifications.append({
                    "type": "overdue_invoice",
                    "severity": "error",
                    "title": "Vadesi Geçmiş Fatura",
                    "message": f"{invoice['invoice_number']} numaralı fatura vadesi geçmiş ({invoice['dealer_name']})",
                    "invoice_id": invoice["id"]
                })
        pending_orders = await db.orders.find({"status": "pending"}, {"_id": 0}).to_list(100)
        for order in pending_orders:
            notifications.append({
                "type": "pending_order",
                "severity": "info",
                "title": "Bekleyen Sipariş",
                "message": f"{order['order_number']} numaralı sipariş onay bekliyor ({order['dealer_name']})",
                "order_id": order["id"]
            })
        return notifications

    # ==================== ÜRETİM GELİŞTİRMELERİ ====================

    @api_router.put("/production/{production_id}/complete")
    async def complete_production(production_id: str, current_user: dict = Depends(get_current_user)):
        production = await db.production.find_one({"id": production_id}, {"_id": 0})
        if not production:
            raise HTTPException(status_code=404, detail="Üretim emri bulunamadı")
        recipe = await db.recipes.find_one({"id": production["recipe_id"]}, {"_id": 0})
        if recipe:
            multiplier = production["quantity"] / recipe.get("yield_quantity", 1)
            for ingredient in recipe.get("ingredients", []):
                required_qty = ingredient["quantity"] * multiplier
                await db.materials.update_one(
                    {"id": ingredient["material_id"]},
                    {"$inc": {"stock_quantity": -required_qty}}
                )
                await db.stock_movements.insert_one({
                    "id": str(uuid.uuid4()),
                    "material_id": ingredient["material_id"],
                    "material_name": ingredient["material_name"],
                    "type": "out",
                    "quantity": required_qty,
                    "reason": f"Üretim: {production['product_name']} ({production_id[:8]})",
                    "reference_id": production_id,
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
        await db.production.update_one({"id": production_id}, {"$set": {"status": "completed", "completed_at": datetime.now(timezone.utc).isoformat()}})
        return {"message": "Üretim tamamlandı ve stoklar güncellendi"}

    @api_router.get("/production/{production_id}/cost")
    async def calculate_production_cost(production_id: str, current_user: dict = Depends(get_current_user)):
        production = await db.production.find_one({"id": production_id}, {"_id": 0})
        if not production:
            raise HTTPException(status_code=404, detail="Üretim emri bulunamadı")
        recipe = await db.recipes.find_one({"id": production["recipe_id"]}, {"_id": 0})
        if not recipe:
            return {"production_id": production_id, "total_cost": 0, "cost_breakdown": []}
        multiplier = production["quantity"] / recipe.get("yield_quantity", 1)
        cost_breakdown = []
        total_cost = 0
        for ingredient in recipe.get("ingredients", []):
            material = await db.materials.find_one({"id": ingredient["material_id"]}, {"_id": 0})
            unit_price = material.get("unit_price", 0) if material else 0
            required_qty = ingredient["quantity"] * multiplier
            cost = required_qty * unit_price
            total_cost += cost
            cost_breakdown.append({
                "material_name": ingredient["material_name"],
                "quantity": required_qty,
                "unit_price": unit_price,
                "cost": cost
            })
        return {
            "production_id": production_id,
            "product_name": production["product_name"],
            "quantity": production["quantity"],
            "total_cost": total_cost,
            "unit_cost": total_cost / production["quantity"] if production["quantity"] > 0 else 0,
            "cost_breakdown": cost_breakdown
        }

    # ==================== EXCEL İMPORT ====================

    from fastapi import UploadFile, File

    @api_router.post("/import/products-excel")
    async def import_products_excel(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
        if not openpyxl_available:
            raise HTTPException(status_code=503, detail="Excel import not available")
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Sadece Excel dosyası (.xlsx) yükleyebilirsiniz")
        try:
            from openpyxl import load_workbook
            contents = await file.read()
            wb = load_workbook(io.BytesIO(contents))
            ws = wb.active
            imported = 0
            errors = []
            headers = [cell.value for cell in ws[1]]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue
                try:
                    product_data = {
                        "id": str(uuid.uuid4()),
                        "code": str(row[0]) if row[0] else f"P-{row_idx}",
                        "name": str(row[1]) if len(row) > 1 and row[1] else f"Ürün {row_idx}",
                        "unit": str(row[2]) if len(row) > 2 and row[2] else "kg",
                        "base_price": float(row[3]) if len(row) > 3 and row[3] else 0,
                        "description": str(row[4]) if len(row) > 4 and row[4] else "",
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    existing = await db.products.find_one({"code": product_data["code"]})
                    if existing:
                        await db.products.update_one({"code": product_data["code"]}, {"$set": {
                            "name": product_data["name"],
                            "unit": product_data["unit"],
                            "base_price": product_data["base_price"],
                            "description": product_data["description"]
                        }})
                    else:
                        await db.products.insert_one(product_data)
                    imported += 1
                except Exception as e:
                    errors.append(f"Satır {row_idx}: {str(e)}")
            return {"imported": imported, "errors": errors, "message": f"{imported} ürün başarıyla içe aktarıldı"}
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Excel okuma hatası: {str(e)}")

    @api_router.post("/import/materials-excel")
    async def import_materials_excel(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
        if not openpyxl_available:
            raise HTTPException(status_code=503, detail="Excel import not available")
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Sadece Excel dosyası (.xlsx) yükleyebilirsiniz")
        try:
            from openpyxl import load_workbook
            contents = await file.read()
            wb = load_workbook(io.BytesIO(contents))
            ws = wb.active
            imported = 0
            errors = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue
                try:
                    material_data = {
                        "id": str(uuid.uuid4()),
                        "code": str(row[0]) if row[0] else f"M-{row_idx}",
                        "name": str(row[1]) if len(row) > 1 and row[1] else f"Hammadde {row_idx}",
                        "unit": str(row[2]) if len(row) > 2 and row[2] else "kg",
                        "stock_quantity": float(row[3]) if len(row) > 3 and row[3] else 0,
                        "min_stock": float(row[4]) if len(row) > 4 and row[4] else 0,
                        "unit_price": float(row[5]) if len(row) > 5 and row[5] else 0,
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    existing = await db.materials.find_one({"code": material_data["code"]})
                    if existing:
                        await db.materials.update_one({"code": material_data["code"]}, {"$set": {
                            "name": material_data["name"],
                            "unit": material_data["unit"],
                            "stock_quantity": material_data["stock_quantity"],
                            "min_stock": material_data["min_stock"],
                            "unit_price": material_data["unit_price"]
                        }})
                    else:
                        await db.materials.insert_one(material_data)
                    imported += 1
                except Exception as e:
                    errors.append(f"Satır {row_idx}: {str(e)}")
            return {"imported": imported, "errors": errors, "message": f"{imported} hammadde başarıyla içe aktarıldı"}
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Excel okuma hatası: {str(e)}")

    @api_router.get("/templates/products-excel")
    async def get_products_template(current_user: dict = Depends(get_current_user)):
        if not openpyxl_available:
            raise HTTPException(status_code=503, detail="Excel not available")
        wb = Workbook()
        ws = wb.active
        ws.title = "Ürünler"
        headers = ["Kod", "Ürün Adı", "Birim", "Fiyat (TL)", "Açıklama"]
        header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        ws.append(["BK-001", "Klasik Burger Köftesi", "kg", "150.00", "200gr porsiyon"])
        ws.append(["BK-002", "Acılı Burger Köftesi", "kg", "160.00", "Acı biberli"])
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 5
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=urun_sablonu.xlsx"})

    @api_router.get("/templates/materials-excel")
    async def get_materials_template(current_user: dict = Depends(get_current_user)):
        if not openpyxl_available:
            raise HTTPException(status_code=503, detail="Excel not available")
        wb = Workbook()
        ws = wb.active
        ws.title = "Hammaddeler"
        headers = ["Kod", "Hammadde Adı", "Birim", "Stok Miktarı", "Min Stok", "Birim Fiyat (TL)"]
        header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        ws.append(["HM-001", "Dana Kıyma", "kg", "100", "20", "250.00"])
        ws.append(["HM-002", "Soğan", "kg", "50", "10", "15.00"])
        ws.append(["HM-003", "Tuz", "kg", "25", "5", "8.00"])
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 5
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=hammadde_sablonu.xlsx"})

    # ==================== COMPANY SETTINGS ====================

    class CompanySettings(BaseModel):
        name: str = "KasaBurger"
        address: Optional[str] = ""
        phone: Optional[str] = ""
        email: Optional[str] = ""
        tax_number: Optional[str] = ""
        tax_office: Optional[str] = ""
        type: Optional[str] = None
        updated_at: Optional[str] = None
        updated_by: Optional[str] = None

    @api_router.get("/settings/company")
    async def get_company_settings(current_user: dict = Depends(get_current_user)):
        settings = await db.settings.find_one({"type": "company"}, {"_id": 0})
        if not settings:
            return CompanySettings().model_dump()
        return settings

    @api_router.put("/settings/company")
    async def update_company_settings(settings: CompanySettings, current_user: dict = Depends(get_current_user)):
        settings_doc = {
            "type": "company",
            **settings.model_dump(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user["name"]
        }
        await db.settings.update_one(
            {"type": "company"},
            {"$set": settings_doc},
            upsert=True
        )
        return {"message": "Şirket bilgileri kaydedildi", "settings": settings_doc}

    # Bildirim Ayarları (NetGSM & SMTP)
    class NotificationSettings(BaseModel):
        netgsm_usercode: Optional[str] = ""
        netgsm_password: Optional[str] = ""
        netgsm_header: Optional[str] = ""
        smtp_host: Optional[str] = ""
        smtp_port: Optional[int] = 587
        smtp_user: Optional[str] = ""
        smtp_password: Optional[str] = ""
        smtp_from: Optional[str] = ""

    @api_router.get("/settings/notifications")
    async def get_notification_settings(current_user: dict = Depends(get_current_user)):
        settings = await db.settings.find_one({"type": "notifications"}, {"_id": 0})
        if not settings:
            return {
                "netgsm_usercode": "",
                "netgsm_password": "",
                "netgsm_header": "",
                "smtp_host": "",
                "smtp_port": 587,
                "smtp_user": "",
                "smtp_password": "",
                "smtp_from": ""
            }
        # Şifreleri maskele
        if settings.get("netgsm_password"):
            settings["netgsm_password"] = "********"
        if settings.get("smtp_password"):
            settings["smtp_password"] = "********"
        return settings

    @api_router.put("/settings/notifications")
    async def update_notification_settings(settings: NotificationSettings, current_user: dict = Depends(get_current_user)):
        # Mevcut ayarları al
        existing = await db.settings.find_one({"type": "notifications"}, {"_id": 0})
        
        settings_doc = {
            "type": "notifications",
            **settings.model_dump(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Şifre ******** ise eski şifreyi koru
        if settings.netgsm_password == "********" and existing:
            settings_doc["netgsm_password"] = existing.get("netgsm_password", "")
        if settings.smtp_password == "********" and existing:
            settings_doc["smtp_password"] = existing.get("smtp_password", "")
        
        await db.settings.update_one(
            {"type": "notifications"},
            {"$set": settings_doc},
            upsert=True
        )
        return {"message": "Bildirim ayarları kaydedildi"}

    @api_router.post("/settings/test-sms")
    async def test_sms_settings(phone: str, current_user: dict = Depends(get_current_user)):
        """SMS ayarlarını test et"""
        settings = await db.settings.find_one({"type": "notifications"}, {"_id": 0})
        if not settings or not settings.get("netgsm_usercode"):
            raise HTTPException(status_code=400, detail="NetGSM ayarları yapılmamış")
        
        # Test SMS gönder
        phone = phone.replace(" ", "").replace("-", "")
        if phone.startswith("0"):
            phone = "90" + phone[1:]
        elif not phone.startswith("90"):
            phone = "90" + phone
        
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                response = await client.get(
                    "https://api.netgsm.com.tr/sms/send/get",
                    params={
                        "usercode": settings["netgsm_usercode"],
                        "password": settings["netgsm_password"],
                        "gsmno": phone,
                        "message": "KasaBurger SMS test mesajı. Bu mesajı aldıysanız ayarlarınız doğru çalışıyor.",
                        "msgheader": settings["netgsm_header"],
                        "dil": "TR"
                    }
                )
                result = response.text.strip()
                if result.startswith("00") or result.startswith("01") or result.startswith("02"):
                    return {"status": "success", "message": "Test SMS gönderildi"}
                else:
                    return {"status": "error", "message": f"NetGSM hatası: {result}"}
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"SMS gönderim hatası: {str(e)}")

    @api_router.post("/settings/test-email")
    async def test_email_settings(email: str, current_user: dict = Depends(get_current_user)):
        """Email ayarlarını test et"""
        import smtplib
        from email.mime.text import MIMEText
        
        settings = await db.settings.find_one({"type": "notifications"}, {"_id": 0})
        if not settings or not settings.get("smtp_host"):
            raise HTTPException(status_code=400, detail="SMTP ayarları yapılmamış")
        
        try:
            server = smtplib.SMTP(settings["smtp_host"], int(settings.get("smtp_port", 587)))
            server.starttls()
            server.login(settings["smtp_user"], settings["smtp_password"])
            
            msg = MIMEText("Bu bir test mesajıdır. Email ayarlarınız doğru çalışıyor.", "plain", "utf-8")
            msg["Subject"] = "KasaBurger - Email Test"
            msg["From"] = settings["smtp_from"]
            msg["To"] = email
            
            server.sendmail(settings["smtp_from"], email, msg.as_string())
            server.quit()
            
            return {"status": "success", "message": "Test email gönderildi"}
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Email gönderim hatası: {str(e)}")

    # ==================== ÖDEME (PAYMENTS) MODÜLÜ ====================

    class PaymentCreate(BaseModel):
        invoice_id: str
        dealer_id: str
        dealer_name: str
        amount: float
        payment_method: str  # cash, bank_transfer, credit_card, check
        payment_date: str
        reference_no: Optional[str] = ""
        notes: Optional[str] = ""

    @api_router.post("/payments")
    async def create_payment(payment: PaymentCreate, current_user: dict = Depends(get_current_user)):
        # Verify invoice exists
        invoice = await db.invoices.find_one({"id": payment.invoice_id}, {"_id": 0})
        if not invoice:
            raise HTTPException(status_code=404, detail="Fatura bulunamadı")
        
        payment_id = str(uuid.uuid4())
        payment_doc = {
            "id": payment_id,
            "payment_number": f"PAY-{datetime.now().strftime('%Y%m%d')}-{payment_id[:4].upper()}",
            **payment.model_dump(),
            "invoice_number": invoice.get("invoice_number", ""),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user["name"]
        }
        await db.payments.insert_one(payment_doc)
        
        # Update invoice paid amount
        current_paid = invoice.get("paid_amount", 0)
        new_paid = current_paid + payment.amount
        invoice_total = invoice.get("total", 0)
        new_status = "paid" if new_paid >= invoice_total else "partial"
        
        await db.invoices.update_one(
            {"id": payment.invoice_id},
            {"$set": {"paid_amount": new_paid, "status": new_status}}
        )
        
        # Update dealer balance
        await db.dealers.update_one(
            {"id": payment.dealer_id},
            {"$inc": {"balance": -payment.amount}}
        )
        
        # Create transaction record
        await db.transactions.insert_one({
            "id": str(uuid.uuid4()),
            "type": "income",
            "category": "Tahsilat",
            "amount": payment.amount,
            "description": f"Fatura ödemesi: {invoice.get('invoice_number', '')} - {payment.dealer_name}",
            "reference_id": payment_id,
            "reference_type": "payment",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user["name"]
        })
        
        return {k: v for k, v in payment_doc.items() if k != "_id"}

    @api_router.get("/payments/summary")
    async def get_payments_summary(current_user: dict = Depends(get_current_user)):
        # Sadece gerekli alanları al - performans optimizasyonu
        payments = await db.payments.find({}, {"_id": 0, "amount": 1, "payment_method": 1}).to_list(10000)
        
        total_collected = sum(p.get("amount", 0) for p in payments)
        
        # Group by payment method
        by_method = {}
        for p in payments:
            method = p.get("payment_method", "other")
            by_method[method] = by_method.get(method, 0) + p.get("amount", 0)
        
        # Get unpaid invoices total - sadece gerekli alanlar
        unpaid_invoices = await db.invoices.find({"status": {"$in": ["unpaid", "partial"]}}, {"_id": 0, "total": 1, "paid_amount": 1}).to_list(10000)
        total_unpaid = sum(i.get("total", 0) - i.get("paid_amount", 0) for i in unpaid_invoices)
        
        return {
            "total_collected": total_collected,
            "total_unpaid": total_unpaid,
            "payment_count": len(payments),
            "by_method": by_method
        }

    @api_router.get("/payments/by-dealer/{dealer_id}")
    async def get_payments_by_dealer(dealer_id: str, current_user: dict = Depends(get_current_user)):
        payments = await db.payments.find({"dealer_id": dealer_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
        return payments

    @api_router.get("/payments/by-invoice/{invoice_id}")
    async def get_payments_by_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
        payments = await db.payments.find({"invoice_id": invoice_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
        return payments

    @api_router.get("/payments")
    async def get_payments(current_user: dict = Depends(get_current_user)):
        payments = await db.payments.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
        return payments

    @api_router.get("/payments/{payment_id}")
    async def get_payment(payment_id: str, current_user: dict = Depends(get_current_user)):
        payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
        if not payment:
            raise HTTPException(status_code=404, detail="Ödeme bulunamadı")
        return payment

    @api_router.delete("/payments/{payment_id}")
    async def delete_payment(payment_id: str, current_user: dict = Depends(get_current_user)):
        payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
        if not payment:
            raise HTTPException(status_code=404, detail="Ödeme bulunamadı")
        
        # Revert invoice paid amount
        await db.invoices.update_one(
            {"id": payment["invoice_id"]},
            {"$inc": {"paid_amount": -payment["amount"]}}
        )
        
        # Check if invoice should be unpaid again
        invoice = await db.invoices.find_one({"id": payment["invoice_id"]}, {"_id": 0})
        if invoice:
            new_paid = invoice.get("paid_amount", 0) - payment["amount"]
            new_status = "unpaid" if new_paid <= 0 else "partial"
            await db.invoices.update_one(
                {"id": payment["invoice_id"]},
                {"$set": {"status": new_status}}
            )
        
        # Revert dealer balance
        await db.dealers.update_one(
            {"id": payment["dealer_id"]},
            {"$inc": {"balance": payment["amount"]}}
        )
        
        # Delete payment
        await db.payments.delete_one({"id": payment_id})
        
        return {"message": "Ödeme silindi"}

    @api_router.get("/payments/by-dealer/{dealer_id}")
    async def get_payments_by_dealer(dealer_id: str, current_user: dict = Depends(get_current_user)):
        payments = await db.payments.find({"dealer_id": dealer_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
        return payments

    @api_router.get("/payments/by-invoice/{invoice_id}")
    async def get_payments_by_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
        payments = await db.payments.find({"invoice_id": invoice_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
        return payments

    # ==================== KAMPANYA YÖNETİMİ ====================
    
    class CampaignCreate(BaseModel):
        title: str
        description: str
        campaign_type: str  # discount, new_product, announcement
        discount_type: Optional[str] = None  # percent, amount
        discount_value: Optional[float] = None
        start_date: str
        end_date: str
        target_dealers: Optional[List[str]] = []  # boş = tüm bayiler
        send_sms: bool = False
        send_email: bool = False

    @api_router.get("/campaigns")
    async def get_campaigns(current_user: dict = Depends(get_current_user)):
        campaigns = await db.campaigns.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
        return campaigns

    @api_router.post("/campaigns")
    async def create_campaign(campaign: CampaignCreate, current_user: dict = Depends(get_current_user)):
        campaign_id = str(uuid.uuid4())
        campaign_doc = {
            "id": campaign_id,
            **campaign.model_dump(),
            "status": "active",
            "sms_sent": False,
            "email_sent": False,
            "sms_count": 0,
            "email_count": 0,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user["name"]
        }
        await db.campaigns.insert_one(campaign_doc)
        
        # Bildirim gönder
        notification_results = {"sms": None, "email": None}
        
        if campaign.send_sms or campaign.send_email:
            # Hedef bayileri al
            if campaign.target_dealers and len(campaign.target_dealers) > 0:
                dealers = await db.dealers.find({"id": {"$in": campaign.target_dealers}}, {"_id": 0}).to_list(100)
            else:
                dealers = await db.dealers.find({}, {"_id": 0}).to_list(100)
            
            if campaign.send_sms:
                sms_result = await send_campaign_sms(campaign_doc, dealers)
                notification_results["sms"] = sms_result
                await db.campaigns.update_one({"id": campaign_id}, {"$set": {"sms_sent": True, "sms_count": sms_result.get("sent", 0)}})
            
            if campaign.send_email:
                email_result = await send_campaign_email(campaign_doc, dealers)
                notification_results["email"] = email_result
                await db.campaigns.update_one({"id": campaign_id}, {"$set": {"email_sent": True, "email_count": email_result.get("sent", 0)}})
        
        return {
            "campaign": {k: v for k, v in campaign_doc.items() if k != "_id"},
            "notifications": notification_results
        }

    @api_router.get("/campaigns/{campaign_id}")
    async def get_campaign(campaign_id: str, current_user: dict = Depends(get_current_user)):
        campaign = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
        if not campaign:
            raise HTTPException(status_code=404, detail="Kampanya bulunamadı")
        return campaign

    @api_router.put("/campaigns/{campaign_id}")
    async def update_campaign(campaign_id: str, campaign: CampaignCreate, current_user: dict = Depends(get_current_user)):
        result = await db.campaigns.update_one({"id": campaign_id}, {"$set": campaign.model_dump()})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Kampanya bulunamadı")
        updated = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
        return updated

    @api_router.delete("/campaigns/{campaign_id}")
    async def delete_campaign(campaign_id: str, current_user: dict = Depends(get_current_user)):
        result = await db.campaigns.delete_one({"id": campaign_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Kampanya bulunamadı")
        return {"message": "Kampanya silindi"}

    @api_router.post("/campaigns/{campaign_id}/send")
    async def send_campaign_notifications(campaign_id: str, send_sms: bool = True, send_email: bool = True, current_user: dict = Depends(get_current_user)):
        """Kampanya bildirimlerini tekrar gönder"""
        campaign = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
        if not campaign:
            raise HTTPException(status_code=404, detail="Kampanya bulunamadı")
        
        # Hedef bayileri al
        target_dealers = campaign.get("target_dealers", [])
        if target_dealers and len(target_dealers) > 0:
            dealers = await db.dealers.find({"id": {"$in": target_dealers}}, {"_id": 0}).to_list(100)
        else:
            dealers = await db.dealers.find({}, {"_id": 0}).to_list(100)
        
        results = {"sms": None, "email": None}
        
        if send_sms:
            results["sms"] = await send_campaign_sms(campaign, dealers)
            await db.campaigns.update_one({"id": campaign_id}, {"$set": {"sms_sent": True, "sms_count": results["sms"].get("sent", 0)}})
        
        if send_email:
            results["email"] = await send_campaign_email(campaign, dealers)
            await db.campaigns.update_one({"id": campaign_id}, {"$set": {"email_sent": True, "email_count": results["email"].get("sent", 0)}})
        
        return results

    # NetGSM SMS Gönderimi
    async def send_campaign_sms(campaign: dict, dealers: list):
        """NetGSM ile SMS gönderir"""
        settings = await db.settings.find_one({"type": "notifications"}, {"_id": 0})
        
        netgsm_usercode = settings.get("netgsm_usercode") if settings else None
        netgsm_password = settings.get("netgsm_password") if settings else None
        netgsm_header = settings.get("netgsm_header") if settings else None
        
        if not all([netgsm_usercode, netgsm_password, netgsm_header]):
            return {"status": "error", "message": "NetGSM ayarları yapılmamış. Ayarlar sayfasından yapılandırın.", "sent": 0}
        
        # Kampanya mesajı oluştur
        message = create_campaign_message(campaign)
        
        sent_count = 0
        errors = []
        
        for dealer in dealers:
            phone = dealer.get("phone", "").replace(" ", "").replace("-", "")
            if not phone:
                continue
            
            # Türkiye formatına çevir
            if phone.startswith("0"):
                phone = "90" + phone[1:]
            elif not phone.startswith("90"):
                phone = "90" + phone
            
            try:
                # NetGSM API çağrısı
                async with httpx.AsyncClient(timeout=30.0) as client:
                    response = await client.get(
                        "https://api.netgsm.com.tr/sms/send/get",
                        params={
                            "usercode": netgsm_usercode,
                            "password": netgsm_password,
                            "gsmno": phone,
                            "message": message,
                            "msgheader": netgsm_header,
                            "dil": "TR"
                        }
                    )
                    
                    result = response.text.strip()
                    if result.startswith("00") or result.startswith("01") or result.startswith("02"):
                        sent_count += 1
                    else:
                        errors.append(f"{dealer.get('name', phone)}: {result}")
                        
            except Exception as e:
                errors.append(f"{dealer.get('name', phone)}: {str(e)}")
        
        return {
            "status": "success" if sent_count > 0 else "error",
            "sent": sent_count,
            "total": len(dealers),
            "errors": errors[:5] if errors else []
        }

    # SMTP Email Gönderimi
    async def send_campaign_email(campaign: dict, dealers: list):
        """SMTP ile email gönderir"""
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        
        settings = await db.settings.find_one({"type": "notifications"}, {"_id": 0})
        
        smtp_host = settings.get("smtp_host") if settings else None
        smtp_port = settings.get("smtp_port", 587) if settings else 587
        smtp_user = settings.get("smtp_user") if settings else None
        smtp_password = settings.get("smtp_password") if settings else None
        smtp_from = settings.get("smtp_from") if settings else None
        
        if not all([smtp_host, smtp_user, smtp_password, smtp_from]):
            return {"status": "error", "message": "SMTP ayarları yapılmamış. Ayarlar sayfasından yapılandırın.", "sent": 0}
        
        # Email içeriği oluştur
        subject, html_body = create_campaign_email(campaign)
        
        sent_count = 0
        errors = []
        
        try:
            server = smtplib.SMTP(smtp_host, int(smtp_port))
            server.starttls()
            server.login(smtp_user, smtp_password)
            
            for dealer in dealers:
                email = dealer.get("email", "")
                if not email or "@" not in email:
                    continue
                
                try:
                    msg = MIMEMultipart("alternative")
                    msg["Subject"] = subject
                    msg["From"] = smtp_from
                    msg["To"] = email
                    
                    html_part = MIMEText(html_body.replace("{dealer_name}", dealer.get("name", "Değerli Bayimiz")), "html", "utf-8")
                    msg.attach(html_part)
                    
                    server.sendmail(smtp_from, email, msg.as_string())
                    sent_count += 1
                    
                except Exception as e:
                    errors.append(f"{dealer.get('name', email)}: {str(e)}")
            
            server.quit()
            
        except Exception as e:
            return {"status": "error", "message": f"SMTP bağlantı hatası: {str(e)}", "sent": 0}
        
        return {
            "status": "success" if sent_count > 0 else "error",
            "sent": sent_count,
            "total": len(dealers),
            "errors": errors[:5] if errors else []
        }

    def create_campaign_message(campaign: dict) -> str:
        """SMS için kampanya mesajı oluşturur"""
        msg = f"KasaBurger: {campaign['title']}\n"
        
        if campaign.get("campaign_type") == "discount":
            if campaign.get("discount_type") == "percent":
                msg += f"%{campaign.get('discount_value', 0)} indirim!\n"
            else:
                msg += f"{campaign.get('discount_value', 0)} TL indirim!\n"
        
        msg += campaign.get("description", "")[:100]
        
        if campaign.get("end_date"):
            msg += f"\nSon tarih: {campaign['end_date'][:10]}"
        
        return msg[:160]  # SMS karakter limiti

    def create_campaign_email(campaign: dict) -> tuple:
        """Email için kampanya içeriği oluşturur"""
        subject = f"KasaBurger - {campaign['title']}"
        
        discount_info = ""
        if campaign.get("campaign_type") == "discount":
            if campaign.get("discount_type") == "percent":
                discount_info = f'<div style="font-size: 32px; color: #f97316; font-weight: bold;">%{campaign.get("discount_value", 0)} İNDİRİM</div>'
            else:
                discount_info = f'<div style="font-size: 32px; color: #f97316; font-weight: bold;">{campaign.get("discount_value", 0)} TL İNDİRİM</div>'
        
        campaign_type_text = {
            "discount": "İndirim Kampanyası",
            "new_product": "Yeni Ürün Duyurusu",
            "announcement": "Duyuru"
        }.get(campaign.get("campaign_type", ""), "Kampanya")
        
        html_body = f'''
        <!DOCTYPE html>
        <html>
        <head><meta charset="UTF-8"></head>
        <body style="font-family: Arial, sans-serif; background-color: #f5f5f5; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 10px; overflow: hidden; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                <div style="background: linear-gradient(135deg, #f97316, #ea580c); padding: 30px; text-align: center;">
                    <h1 style="color: white; margin: 0; font-size: 28px;">KasaBurger</h1>
                    <p style="color: rgba(255,255,255,0.9); margin: 10px 0 0 0;">{campaign_type_text}</p>
                </div>
                <div style="padding: 30px; text-align: center;">
                    <h2 style="color: #333; margin: 0 0 20px 0;">{campaign['title']}</h2>
                    {discount_info}
                    <p style="color: #666; line-height: 1.6; margin: 20px 0;">{{dealer_name}},</p>
                    <p style="color: #666; line-height: 1.6;">{campaign.get('description', '')}</p>
                    <div style="margin: 30px 0; padding: 15px; background: #fff3e0; border-radius: 8px;">
                        <p style="margin: 0; color: #e65100;">
                            <strong>Kampanya Tarihleri:</strong><br>
                            {campaign.get('start_date', '')[:10]} - {campaign.get('end_date', '')[:10]}
                        </p>
                    </div>
                    <a href="https://erp.kasaburger.net.tr/dealer-login" style="display: inline-block; background: #f97316; color: white; padding: 15px 30px; text-decoration: none; border-radius: 8px; font-weight: bold;">Sipariş Ver</a>
                </div>
                <div style="background: #f5f5f5; padding: 20px; text-align: center; color: #999; font-size: 12px;">
                    <p>Bu email KasaBurger tarafından gönderilmiştir.</p>
                </div>
            </div>
        </body>
        </html>
        '''
        
        return subject, html_body

    # ==================== API ROOT ====================

    @api_router.get("/")
    async def api_root():
        return {"message": "KasaBurger API v1.0.3", "status": "active", "security": "enabled"}

    @api_router.get("/health")
    async def api_health():
        return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

    @api_router.get("/security/status")
    async def security_status(current_user: dict = Depends(get_current_user)):
        """Admin için güvenlik durumu - sadece admin görebilir"""
        if current_user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Yetkiniz yok")
        
        return {
            "status": "active",
            "features": {
                "rate_limiting": True,
                "brute_force_protection": True,
                "cors_restricted": True,
                "security_headers": True,
                "ip_blocking": True,
                "captcha": True,
                "two_factor_auth": True,
                "password_policy": True,
                "audit_logging": True
            },
            "settings": {
                "rate_limit_requests": RATE_LIMIT_REQUESTS,
                "rate_limit_window_seconds": RATE_LIMIT_WINDOW,
                "login_attempt_limit": LOGIN_ATTEMPT_LIMIT,
                "login_block_duration_seconds": LOGIN_BLOCK_DURATION,
                "bulk_request_limit": BULK_REQUEST_LIMIT,
                "password_min_length": PASSWORD_MIN_LENGTH,
                "captcha_after_failed_attempts": 2
            },
            "current_stats": {
                "blocked_ips_count": len(blocked_ips),
                "tracked_ips_count": len(rate_limit_storage),
                "pending_2fa_codes": len(two_factor_codes),
                "audit_log_entries": len(audit_logs)
            }
        }

    @api_router.get("/security/audit-logs")
    async def get_audit_logs(current_user: dict = Depends(get_current_user), limit: int = 50):
        """Admin için audit loglarını getir"""
        if current_user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Yetkiniz yok")
        
        # Son logları döndür (en yeniden en eskiye)
        return audit_logs[-limit:][::-1]

    @api_router.post("/security/unblock-ip")
    async def unblock_ip(ip_address: str, current_user: dict = Depends(get_current_user)):
        """Admin için IP engelini kaldır"""
        if current_user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Yetkiniz yok")
        
        if ip_address in blocked_ips:
            del blocked_ips[ip_address]
            log_audit("IP_UNBLOCKED", current_user["id"], current_user["email"], "admin", ip_address)
            return {"message": f"IP adresi engeli kaldırıldı: {ip_address}"}
        
        # Check dealer blocked IPs too
        dealer_key = f"dealer_{ip_address}"
        if dealer_key in blocked_ips:
            del blocked_ips[dealer_key]
            log_audit("IP_UNBLOCKED", current_user["id"], current_user["email"], "admin", ip_address)
            return {"message": f"Bayi IP adresi engeli kaldırıldı: {ip_address}"}
        
        return {"message": "Bu IP adresi engelli değil"}

    @api_router.get("/security/password-policy")
    async def get_password_policy():
        """Şifre politikasını getir"""
        return {
            "min_length": PASSWORD_MIN_LENGTH,
            "require_uppercase": PASSWORD_REQUIRE_UPPERCASE,
            "require_lowercase": PASSWORD_REQUIRE_LOWERCASE,
            "require_number": PASSWORD_REQUIRE_NUMBER,
            "require_special": PASSWORD_REQUIRE_SPECIAL
        }

    # ==================== KIOSK YÖNETİMİ ====================
    
    class KioskProduct(BaseModel):
        id: str
        name: str
        description: Optional[str] = ""
        price: float
        category: str
        image: Optional[str] = ""
        premium: Optional[bool] = False

    class KioskOrder(BaseModel):
        items: List[dict]
        total: float
        service_type: str  # 'paket' or 'masa'
        table_number: Optional[str] = None
        payment_method: str  # 'cash' or 'card'

    @api_router.get("/kiosk/menu")
    async def get_kiosk_menu():
        """Kiosk menüsünü getir - Auth gerekmez"""
        products = await db.kiosk_products.find({}, {"_id": 0}).to_list(1000)
        
        categories = [
            {"id": "et-burger", "name": "Et Burger", "icon": "🍔"},
            {"id": "premium", "name": "Premium Gourmet", "icon": "👑"},
            {"id": "tavuk", "name": "Tavuk Burger", "icon": "🍗"},
            {"id": "atistirmalik", "name": "Atıştırmalıklar", "icon": "🍟"},
            {"id": "icecek", "name": "İçecekler", "icon": "🥤"},
            {"id": "tatli", "name": "Tatlılar", "icon": "🍫"},
        ]
        
        return {"categories": categories, "products": products}

    @api_router.get("/kiosk/products")
    async def get_kiosk_products(current_user: dict = Depends(get_current_user)):
        """Admin için kiosk ürünlerini listele"""
        products = await db.kiosk_products.find({}, {"_id": 0}).to_list(1000)
        return products

    @api_router.post("/kiosk/products/seed")
    async def seed_kiosk_products(current_user: dict = Depends(get_current_user)):
        """Varsayılan kiosk ürünlerini veritabanına ekle - SADECE BOŞ İSE"""
        existing_count = await db.kiosk_products.count_documents({})
        if existing_count > 0:
            return {"message": f"Veritabanında zaten {existing_count} ürün var. Değişiklik yapılmadı.", "seeded": False, "count": existing_count}
        
        default_products = [
            # ET BURGER
            {"id": "kasa-classic", "name": "Kasa Classic Burger", "category": "et-burger", "price": 460, "image": "https://images.unsplash.com/photo-1599082267955-266a170c214e?w=600&q=80", "is_active": True, "description": "150 gr. özel baharatlı dana köfte, taze yeşillik, Kasa Gizli Sos"},
            {"id": "golden-burger", "name": "Golden Burger", "category": "et-burger", "price": 1190, "image": "https://images.unsplash.com/photo-1603508102983-99b101395d1a?w=600&q=80", "is_active": True, "is_premium": True, "description": "150 gr. Dry-Aged köfte, brioche ekmek, yenilebilir altın kaplama, trüf sos, double cheddar"},
            {"id": "cheese-lover", "name": "Cheese Lover Burger", "category": "et-burger", "price": 560, "image": "https://images.unsplash.com/photo-1585238341710-4d3ff484184d?w=600&q=80", "is_active": True, "description": "150 gr. dana köfte, çift cheddar + erimiş peynir sosu, karamelize soğan"},
            {"id": "no7-burger", "name": "No:7 Burger", "category": "et-burger", "price": 540, "image": "https://images.unsplash.com/photo-1625813506062-0aeb1d7a094b?w=600&q=80", "is_active": True, "description": "150 gr. dana köfte, double cheddar, jalapeno, acılı kasa sos, çıtır soğan"},
            {"id": "hirsiz-polis", "name": "Hırsız & Polis Burger", "category": "et-burger", "price": 490, "image": "https://images.unsplash.com/photo-1568901346375-23c9450c58cd?w=600&q=80", "is_active": True, "description": "2x150 gr. dana köfte, Polis sos (tatlı), Hırsız (acı), cheddar"},
            # PREMIUM GOURMET
            {"id": "viking-burger", "name": "Viking Burger", "category": "premium", "price": 430, "image": "https://images.unsplash.com/photo-1553979459-d2229ba7433b?w=600&q=80", "is_active": True, "is_premium": True, "description": "150 gr. dana köfte, 20 gr. cheddar peyniri, çıtır soğan, viking sos"},
            {"id": "milano-burger", "name": "Milano Burger", "category": "premium", "price": 440, "image": "https://images.unsplash.com/photo-1594212699903-ec8a3eca50f5?w=600&q=80", "is_active": True, "is_premium": True, "description": "150gr. dana köfte, mozzarella, kuru domates, pesto mayo, roka"},
            {"id": "kasa-double-xl", "name": "Kasa Double XL", "category": "premium", "price": 640, "image": "https://images.unsplash.com/photo-1551782450-a2132b4ba21d?w=600&q=80", "is_active": True, "is_premium": True, "description": "300 gr. dana köfte, 40 gr. cheddar, karamelize soğan, kasa özel sos"},
            {"id": "smoky-bbq", "name": "Smoky BBQ Burger", "category": "premium", "price": 560, "image": "https://images.unsplash.com/photo-1572802419224-296b0aeee0d9?w=600&q=80", "is_active": True, "is_premium": True, "description": "150 gr. dana köfte, 20 gr. cheddar, kızartılmış pastırma, bbq sos"},
            {"id": "animal-style", "name": "Animal Style Burger", "category": "premium", "price": 550, "image": "https://images.unsplash.com/photo-1586190848861-99aa4a171e90?w=600&q=80", "is_active": True, "is_premium": True, "description": "150 gr. dana köfte, cheddar peynir, karamelize soğan, animal sos"},
            # TAVUK BURGER
            {"id": "crispy-chicken", "name": "Crispy Chicken Burger", "category": "tavuk", "price": 360, "image": "https://images.unsplash.com/photo-1626082927389-6cd097cdc6ec?w=600&q=80", "is_active": True, "description": "Çıtır paneli tavuk göğsü, taze yeşillik, turşu, mayonez"},
            {"id": "double-crispy", "name": "Double Crispy Chicken", "category": "tavuk", "price": 410, "image": "https://images.unsplash.com/photo-1562967914-608f82629710?w=600&q=80", "is_active": True, "description": "Double tavuk, cheddar, taze yeşillik, acılı kasa sos, turşu"},
            {"id": "animal-chicken", "name": "Animal Style Chicken", "category": "tavuk", "price": 430, "image": "https://images.unsplash.com/photo-1606755962773-d324e0a13086?w=600&q=80", "is_active": True, "description": "Çıtır paneli tavuk göğsü, karamelize soğan, double cheddar, animal sos"},
            {"id": "spicy-hirsiz", "name": "(Spicy) Hırsız Burger", "category": "tavuk", "price": 420, "image": "https://images.unsplash.com/photo-1513185158878-8d8c2a2a3da3?w=600&q=80", "is_active": True, "description": "Acı marinasyonlu çıtır tavuk, cheddar, acılı kasa mayonez, jalapeno"},
            {"id": "sweet-polis", "name": "(Sweet) Polis Burger", "category": "tavuk", "price": 420, "image": "https://images.unsplash.com/photo-1610440042657-612c34d95e9f?w=600&q=80", "is_active": True, "description": "Tatlı marinasyonlu çıtır tavuk, tatlı kasa sos, taze yeşillik, mozzarella"},
            {"id": "milano-chicken", "name": "Milano Chicken Burger", "category": "tavuk", "price": 440, "image": "https://images.unsplash.com/photo-1585325701165-351af916e581?w=600&q=80", "is_active": True, "description": "İnce paneli çıtır tavuk, pesto mayo, kurutulmuş domates, mozzarella"},
            {"id": "viking-chicken", "name": "Viking Chicken Burger", "category": "tavuk", "price": 430, "image": "https://images.unsplash.com/photo-1587132137056-bfbf0166836e?w=600&q=80", "is_active": True, "description": "Viking sos, çıtır tavuk, cheddar, kornişon turşu, çıtır soğan"},
            # ATISTIRMALIKLAR
            {"id": "mac-cheese", "name": "Mac and Cheese Topları", "category": "atistirmalik", "price": 170, "image": "https://images.unsplash.com/photo-1543339494-b4cd4f7ba686?w=600&q=80", "is_active": True},
            {"id": "mozarella-sticks", "name": "Mozarella Sticks", "category": "atistirmalik", "price": 210, "image": "https://images.unsplash.com/photo-1548340748-6d2b7d7da280?w=600&q=80", "is_active": True, "description": "6 adet (yarım porsiyon patates ile)"},
            {"id": "sogan-halkasi", "name": "Soğan Halkası", "category": "atistirmalik", "price": 180, "image": "https://images.unsplash.com/photo-1639024471283-03518883512d?w=600&q=80", "is_active": True, "description": "8 adet (yarım porsiyon patates ile)"},
            {"id": "cheese-fries", "name": "Prison Cheese Lover Fries", "category": "atistirmalik", "price": 150, "image": "https://images.unsplash.com/photo-1573080496219-bb080dd4f877?w=600&q=80", "is_active": True, "description": "Cheddar soslu patates"},
            {"id": "truffle-fries", "name": "Prison Truffle Fries", "category": "atistirmalik", "price": 175, "image": "https://images.unsplash.com/photo-1630384060421-cb20aff8a689?w=600&q=80", "is_active": True, "description": "Trüf soslu patates"},
            {"id": "cajun-fries", "name": "Prison Hot Lockdown Fries", "category": "atistirmalik", "price": 160, "image": "https://images.unsplash.com/photo-1598679253544-2c97992403ea?w=600&q=80", "is_active": True, "description": "Cajun baharatlı patates"},
            # İÇECEKLER
            {"id": "ayran", "name": "Ayran", "category": "icecek", "price": 35, "image": "https://images.unsplash.com/photo-1596151163116-98a5033814c2?w=600&q=80", "is_active": True},
            {"id": "su", "name": "Su", "category": "icecek", "price": 20, "image": "https://images.unsplash.com/photo-1564419320461-6870880221ad?w=600&q=80", "is_active": True},
            {"id": "limonata", "name": "Limonata", "category": "icecek", "price": 55, "image": "https://images.unsplash.com/photo-1621263764928-df1444c5e859?w=600&q=80", "is_active": True},
            {"id": "pepsi", "name": "Pepsi", "category": "icecek", "price": 45, "image": "https://images.unsplash.com/photo-1629203851122-3726ecdf080e?w=600&q=80", "is_active": True},
            {"id": "milkshake", "name": "Milkshake", "category": "icecek", "price": 85, "image": "https://images.unsplash.com/photo-1572490122747-3968b75cc699?w=600&q=80", "is_active": True},
            # TATLILAR
            {"id": "choco-bomb", "name": "Kasa Choco Bomb", "category": "tatli", "price": 200, "image": "https://images.unsplash.com/photo-1606313564200-e75d5e30476c?w=600&q=80", "is_active": True},
            {"id": "churros", "name": "Churros", "category": "tatli", "price": 180, "image": "https://images.pexels.com/photos/2035706/pexels-photo-2035706.jpeg?auto=compress&cs=tinysrgb&w=600", "is_active": True},
            {"id": "oreo-dream", "name": "Oreo Dream Cup", "category": "tatli", "price": 220, "image": "https://images.unsplash.com/photo-1612078960206-1709f1f0c969?w=600&q=80", "is_active": True},
        ]
        
        await db.kiosk_products.insert_many(default_products)
        return {"message": f"{len(default_products)} ürün eklendi", "seeded": True, "count": len(default_products)}

    @api_router.post("/kiosk/products/reset")
    async def reset_kiosk_products(current_user: dict = Depends(get_current_user)):
        """Tüm kiosk ürünlerini sil ve varsayılanları yükle"""
        # Mevcut ürünleri sil
        await db.kiosk_products.delete_many({})
        
        default_products = [
            # ET BURGER
            {"id": "kasa-classic", "name": "Kasa Classic Burger", "category": "et-burger", "price": 460, "image": "https://images.unsplash.com/photo-1599082267955-266a170c214e?w=600&q=80", "is_active": True, "description": "150 gr. özel baharatlı dana köfte, taze yeşillik, Kasa Gizli Sos"},
            {"id": "golden-burger", "name": "Golden Burger", "category": "et-burger", "price": 1190, "image": "https://images.unsplash.com/photo-1603508102983-99b101395d1a?w=600&q=80", "is_active": True, "is_premium": True, "description": "150 gr. Dry-Aged köfte, brioche ekmek, yenilebilir altın kaplama, trüf sos, double cheddar"},
            {"id": "cheese-lover", "name": "Cheese Lover Burger", "category": "et-burger", "price": 560, "image": "https://images.unsplash.com/photo-1585238341710-4d3ff484184d?w=600&q=80", "is_active": True, "description": "150 gr. dana köfte, çift cheddar + erimiş peynir sosu, karamelize soğan"},
            {"id": "no7-burger", "name": "No:7 Burger", "category": "et-burger", "price": 540, "image": "https://images.unsplash.com/photo-1625813506062-0aeb1d7a094b?w=600&q=80", "is_active": True, "description": "150 gr. dana köfte, double cheddar, jalapeno, acılı kasa sos, çıtır soğan"},
            {"id": "hirsiz-polis", "name": "Hırsız & Polis Burger", "category": "et-burger", "price": 490, "image": "https://images.unsplash.com/photo-1568901346375-23c9450c58cd?w=600&q=80", "is_active": True, "description": "2x150 gr. dana köfte, Polis sos (tatlı), Hırsız (acı), cheddar"},
            # PREMIUM GOURMET
            {"id": "viking-burger", "name": "Viking Burger", "category": "premium", "price": 430, "image": "https://images.unsplash.com/photo-1553979459-d2229ba7433b?w=600&q=80", "is_active": True, "is_premium": True, "description": "150 gr. dana köfte, 20 gr. cheddar peyniri, çıtır soğan, viking sos"},
            {"id": "milano-burger", "name": "Milano Burger", "category": "premium", "price": 440, "image": "https://images.unsplash.com/photo-1594212699903-ec8a3eca50f5?w=600&q=80", "is_active": True, "is_premium": True, "description": "150gr. dana köfte, mozzarella, kuru domates, pesto mayo, roka"},
            {"id": "kasa-double-xl", "name": "Kasa Double XL", "category": "premium", "price": 640, "image": "https://images.unsplash.com/photo-1551782450-a2132b4ba21d?w=600&q=80", "is_active": True, "is_premium": True, "description": "300 gr. dana köfte, 40 gr. cheddar, karamelize soğan, kasa özel sos"},
            {"id": "smoky-bbq", "name": "Smoky BBQ Burger", "category": "premium", "price": 560, "image": "https://images.unsplash.com/photo-1572802419224-296b0aeee0d9?w=600&q=80", "is_active": True, "is_premium": True, "description": "150 gr. dana köfte, 20 gr. cheddar, kızartılmış pastırma, bbq sos"},
            {"id": "animal-style", "name": "Animal Style Burger", "category": "premium", "price": 550, "image": "https://images.unsplash.com/photo-1586190848861-99aa4a171e90?w=600&q=80", "is_active": True, "is_premium": True, "description": "150 gr. dana köfte, cheddar peynir, karamelize soğan, animal sos"},
            # TAVUK BURGER
            {"id": "crispy-chicken", "name": "Crispy Chicken Burger", "category": "tavuk", "price": 360, "image": "https://images.unsplash.com/photo-1626082927389-6cd097cdc6ec?w=600&q=80", "is_active": True, "description": "Çıtır paneli tavuk göğsü, taze yeşillik, turşu, mayonez"},
            {"id": "double-crispy", "name": "Double Crispy Chicken", "category": "tavuk", "price": 410, "image": "https://images.unsplash.com/photo-1562967914-608f82629710?w=600&q=80", "is_active": True, "description": "Double tavuk, cheddar, taze yeşillik, acılı kasa sos, turşu"},
            {"id": "animal-chicken", "name": "Animal Style Chicken", "category": "tavuk", "price": 430, "image": "https://images.unsplash.com/photo-1606755962773-d324e0a13086?w=600&q=80", "is_active": True, "description": "Çıtır paneli tavuk göğsü, karamelize soğan, double cheddar, animal sos"},
            {"id": "spicy-hirsiz", "name": "(Spicy) Hırsız Burger", "category": "tavuk", "price": 420, "image": "https://images.unsplash.com/photo-1513185158878-8d8c2a2a3da3?w=600&q=80", "is_active": True, "description": "Acı marinasyonlu çıtır tavuk, cheddar, acılı kasa mayonez, jalapeno"},
            {"id": "sweet-polis", "name": "(Sweet) Polis Burger", "category": "tavuk", "price": 420, "image": "https://images.unsplash.com/photo-1610440042657-612c34d95e9f?w=600&q=80", "is_active": True, "description": "Tatlı marinasyonlu çıtır tavuk, tatlı kasa sos, taze yeşillik, mozzarella"},
            {"id": "milano-chicken", "name": "Milano Chicken Burger", "category": "tavuk", "price": 440, "image": "https://images.unsplash.com/photo-1585325701165-351af916e581?w=600&q=80", "is_active": True, "description": "İnce paneli çıtır tavuk, pesto mayo, kurutulmuş domates, mozzarella"},
            {"id": "viking-chicken", "name": "Viking Chicken Burger", "category": "tavuk", "price": 430, "image": "https://images.unsplash.com/photo-1587132137056-bfbf0166836e?w=600&q=80", "is_active": True, "description": "Viking sos, çıtır tavuk, cheddar, kornişon turşu, çıtır soğan"},
            # ATISTIRMALIKLAR
            {"id": "mac-cheese", "name": "Mac and Cheese Topları", "category": "atistirmalik", "price": 170, "image": "https://images.unsplash.com/photo-1543339494-b4cd4f7ba686?w=600&q=80", "is_active": True},
            {"id": "mozarella-sticks", "name": "Mozarella Sticks", "category": "atistirmalik", "price": 210, "image": "https://images.unsplash.com/photo-1548340748-6d2b7d7da280?w=600&q=80", "is_active": True, "description": "6 adet (yarım porsiyon patates ile)"},
            {"id": "sogan-halkasi", "name": "Soğan Halkası", "category": "atistirmalik", "price": 180, "image": "https://images.unsplash.com/photo-1639024471283-03518883512d?w=600&q=80", "is_active": True, "description": "8 adet (yarım porsiyon patates ile)"},
            {"id": "cheese-fries", "name": "Prison Cheese Lover Fries", "category": "atistirmalik", "price": 150, "image": "https://images.unsplash.com/photo-1573080496219-bb080dd4f877?w=600&q=80", "is_active": True, "description": "Cheddar soslu patates"},
            {"id": "truffle-fries", "name": "Prison Truffle Fries", "category": "atistirmalik", "price": 175, "image": "https://images.unsplash.com/photo-1630384060421-cb20aff8a689?w=600&q=80", "is_active": True, "description": "Trüf soslu patates"},
            {"id": "cajun-fries", "name": "Prison Hot Lockdown Fries", "category": "atistirmalik", "price": 160, "image": "https://images.unsplash.com/photo-1598679253544-2c97992403ea?w=600&q=80", "is_active": True, "description": "Cajun baharatlı patates"},
            # İÇECEKLER
            {"id": "ayran", "name": "Ayran", "category": "icecek", "price": 35, "image": "https://images.unsplash.com/photo-1596151163116-98a5033814c2?w=600&q=80", "is_active": True},
            {"id": "su", "name": "Su", "category": "icecek", "price": 20, "image": "https://images.unsplash.com/photo-1564419320461-6870880221ad?w=600&q=80", "is_active": True},
            {"id": "limonata", "name": "Limonata", "category": "icecek", "price": 55, "image": "https://images.unsplash.com/photo-1621263764928-df1444c5e859?w=600&q=80", "is_active": True},
            {"id": "pepsi", "name": "Pepsi", "category": "icecek", "price": 45, "image": "https://images.unsplash.com/photo-1629203851122-3726ecdf080e?w=600&q=80", "is_active": True},
            {"id": "milkshake", "name": "Milkshake", "category": "icecek", "price": 85, "image": "https://images.unsplash.com/photo-1572490122747-3968b75cc699?w=600&q=80", "is_active": True},
            # TATLILAR
            {"id": "choco-bomb", "name": "Kasa Choco Bomb", "category": "tatli", "price": 200, "image": "https://images.unsplash.com/photo-1606313564200-e75d5e30476c?w=600&q=80", "is_active": True},
            {"id": "churros", "name": "Churros", "category": "tatli", "price": 180, "image": "https://images.pexels.com/photos/2035706/pexels-photo-2035706.jpeg?auto=compress&cs=tinysrgb&w=600", "is_active": True},
            {"id": "oreo-dream", "name": "Oreo Dream Cup", "category": "tatli", "price": 220, "image": "https://images.unsplash.com/photo-1612078960206-1709f1f0c969?w=600&q=80", "is_active": True},
        ]
        
        await db.kiosk_products.insert_many(default_products)
        return {"message": f"Tüm ürünler sıfırlandı. {len(default_products)} ürün yüklendi.", "reset": True, "count": len(default_products)}

    @api_router.post("/kiosk/products")
    async def create_kiosk_product(product: KioskProduct, current_user: dict = Depends(get_current_user)):
        """Admin için yeni kiosk ürünü ekle"""
        existing = await db.kiosk_products.find_one({"id": product.id})
        if existing:
            raise HTTPException(status_code=400, detail="Bu ID ile ürün zaten var")
        
        product_data = product.dict()
        product_data["created_at"] = datetime.now(timezone.utc).isoformat()
        product_data["created_by"] = current_user["email"]
        
        await db.kiosk_products.insert_one(product_data)
        return {"message": "Ürün eklendi", "id": product.id}

    @api_router.put("/kiosk/products/{product_id}")
    async def update_kiosk_product(product_id: str, product: KioskProduct, current_user: dict = Depends(get_current_user)):
        """Admin için kiosk ürünü güncelle"""
        existing = await db.kiosk_products.find_one({"id": product_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        
        product_data = product.dict()
        product_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        product_data["updated_by"] = current_user["email"]
        
        await db.kiosk_products.update_one({"id": product_id}, {"$set": product_data})
        return {"message": "Ürün güncellendi"}

    @api_router.delete("/kiosk/products/{product_id}")
    async def delete_kiosk_product(product_id: str, current_user: dict = Depends(get_current_user)):
        """Admin için kiosk ürünü sil"""
        result = await db.kiosk_products.delete_one({"id": product_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        return {"message": "Ürün silindi"}

    @api_router.post("/kiosk/orders")
    async def create_kiosk_order(order: KioskOrder):
        """Kiosk siparişi oluştur - Auth gerekmez"""
        today = datetime.now().strftime('%Y-%m-%d')
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        
        # Paket ve Masa siparişleri için ayrı numaralama
        if order.service_type == 'paket':
            # Bugünkü paket sipariş sayısını al
            paket_count = await db.kiosk_orders.count_documents({
                "order_date": today,
                "service_type": "paket"
            })
            order_number = f"PKT{str(paket_count + 1).zfill(4)}"
        else:
            # Bugünkü masa sipariş sayısını al
            masa_count = await db.kiosk_orders.count_documents({
                "order_date": today,
                "service_type": "masa"
            })
            order_number = str(masa_count + 1).zfill(4)
        
        order_data = {
            "id": str(uuid.uuid4()),
            "order_number": order_number,
            "order_date": today,
            "items": order.items,
            "total": order.total,
            "service_type": order.service_type,
            "table_number": order.table_number,
            "payment_method": order.payment_method,
            "payment_status": order.payment_method if order.payment_method in ["online", "card_online"] else "pending",
            "status": "new",
            "source": "mobile_qr",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.kiosk_orders.insert_one(order_data)
        return {"order_number": order_number, "id": order_data["id"]}

    @api_router.get("/kiosk/orders")
    async def get_kiosk_orders(current_user: dict = Depends(get_current_user)):
        """Admin için kiosk siparişlerini listele"""
        orders = await db.kiosk_orders.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
        return orders

    @api_router.get("/kiosk/orders/new")
    async def get_new_kiosk_orders(current_user: dict = Depends(get_current_user)):
        """Yeni siparişleri getir (bildirim için)"""
        orders = await db.kiosk_orders.find({"status": "new"}, {"_id": 0}).sort("created_at", -1).to_list(50)
        return orders

    @api_router.put("/kiosk/orders/{order_id}/status")
    async def update_kiosk_order_status(order_id: str, status: str, current_user: dict = Depends(get_current_user)):
        """Sipariş durumunu güncelle"""
        valid_statuses = ["new", "preparing", "ready", "completed", "cancelled"]
        if status not in valid_statuses:
            raise HTTPException(status_code=400, detail=f"Geçersiz durum. Geçerli: {valid_statuses}")
        
        result = await db.kiosk_orders.update_one(
            {"id": order_id},
            {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
        return {"message": "Durum güncellendi"}

    @api_router.get("/kiosk/orders/{order_id}/print")
    async def get_order_print_data(order_id: str, current_user: dict = Depends(get_current_user)):
        """Yazıcı için sipariş verisi"""
        order = await db.kiosk_orders.find_one({"id": order_id}, {"_id": 0})
        if not order:
            raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
        
        # ESC/POS format için hazırla
        print_data = {
            "order_number": order.get("order_number"),
            "service_type": order.get("service_type"),
            "table_number": order.get("table_number"),
            "items": order.get("items", []),
            "total": order.get("total"),
            "created_at": order.get("created_at"),
            "payment_method": order.get("payment_method")
        }
        return print_data

    @api_router.get("/kiosk/qr-code")
    async def generate_qr_code():
        """QR kod oluştur"""
        import qrcode
        import base64
        from io import BytesIO
        
        # Kiosk URL'i
        kiosk_url = "https://erp.kasaburger.net.tr/kiosk"
        
        # QR kod oluştur
        qr = qrcode.QRCode(version=1, box_size=10, border=2)
        qr.add_data(kiosk_url)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Base64'e çevir
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        img_str = base64.b64encode(buffer.getvalue()).decode()
        
        return {
            "qr_code": f"data:image/png;base64,{img_str}",
            "url": kiosk_url
        }

    # ===================== RECIPES ENDPOINTS =====================
    @api_router.get("/recipes")
    async def get_recipes(current_user: dict = Depends(get_current_user)):
        """Tüm reçeteleri listele"""
        recipes = await db.recipes.find({}, {"_id": 0}).to_list(1000)
        return recipes

    @api_router.post("/recipes")
    async def create_recipe(recipe: dict = Body(...), current_user: dict = Depends(get_current_user)):
        """Yeni reçete ekle"""
        recipe["id"] = str(uuid.uuid4())[:8]
        recipe["created_at"] = datetime.now(timezone.utc).isoformat()
        recipe["created_by"] = current_user["email"]
        await db.recipes.insert_one(recipe)
        return {"message": "Reçete eklendi", "id": recipe["id"]}

    @api_router.put("/recipes/{recipe_id}")
    async def update_recipe(recipe_id: str, recipe: dict = Body(...), current_user: dict = Depends(get_current_user)):
        """Reçete güncelle"""
        result = await db.recipes.update_one(
            {"id": recipe_id},
            {"$set": recipe}
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Reçete bulunamadı")
        return {"message": "Reçete güncellendi"}

    @api_router.delete("/recipes/{recipe_id}")
    async def delete_recipe(recipe_id: str, current_user: dict = Depends(get_current_user)):
        """Reçete sil"""
        result = await db.recipes.delete_one({"id": recipe_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Reçete bulunamadı")
        return {"message": "Reçete silindi"}

    # ===================== IMAGE UPLOAD ENDPOINT =====================
    import os
    import shutil
    from fastapi import UploadFile, File
    
    # Uploads klasörünü oluştur
    # Cloudinary'ye resim yükleme endpoint'i
    @api_router.post("/upload/image")
    async def upload_image(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
        """Resim dosyasını Cloudinary'ye yükle"""
        # Dosya tipi kontrolü
        if not file.content_type.startswith('image/'):
            raise HTTPException(status_code=400, detail="Sadece resim dosyaları yüklenebilir")
        
        # Dosya boyutu kontrolü (5MB)
        contents = await file.read()
        if len(contents) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Dosya boyutu 5MB'dan büyük olamaz")
        
        try:
            # Cloudinary'ye yükle
            result = cloudinary.uploader.upload(
                contents,
                folder="kasaburger/products",
                resource_type="image",
                transformation=[
                    {"width": 800, "height": 600, "crop": "limit", "quality": "auto:good"}
                ]
            )
            
            # Cloudinary secure URL'ini döndür
            image_url = result.get("secure_url")
            public_id = result.get("public_id")
            
            logging.info(f"Cloudinary upload successful: {public_id}")
            return {"url": image_url, "public_id": public_id}
            
        except Exception as e:
            logging.error(f"Cloudinary upload error: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Resim yükleme hatası: {str(e)}")

    # Include router
    app.include_router(api_router)

    # Shutdown
    @app.on_event("shutdown")
    async def shutdown_db_client():
        client.close()

    logging.basicConfig(level=logging.INFO)
    logging.info("KasaBurger API v1.0.3 loaded successfully")

except Exception as e:
    # If anything fails, keep health check working
    import logging
    logging.error(f"Failed to load full application: {e}")
    
    @app.get("/api/")
    def api_fallback():
        return {"status": "limited", "error": str(e)}
